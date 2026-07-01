#!/usr/bin/env python3
r"""
CAPI assignment generator — turn the filled assignment-master into the three
artifacts the supervisor actually hands out.

Reads `assignment-master.xlsx` (the "Assignments" sheet) — or a CSV with the
same columns — and writes into ./out/:

  • AS_<enumerator_id>.dat   one fixed-width file per enumerator, matching the
                             hub's ASSIGNMENT_DICT, ready for "Assign Enumeration
                             Area" to serve over Bluetooth (97-byte records).
  • assignment-sheets.html   printable case-key sheets, one page per enumerator
                             (the reliable hand-out path — they type the keys).
  • case-keys.csv            every 12-digit case key, for QA / CSWeb cross-check.
  • manifest.txt             per-enumerator + grand totals, and what each file is for.

The case key = the 9-digit EA/facility code + a 3-digit sequence (001..target).
Use real PSGC codes; a wrong prefix is hard-rejected on the tablet.

    python generate_assignments.py [--master assignment-master.xlsx] [--out out]

Generator-first: edit the master spreadsheet (or this script), never the .dat
files by hand.
"""

from __future__ import annotations

import argparse
import csv
import io
import sys
from html import escape
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)

# ASSIGNMENT_DICT fixed-width layout (relative positions; total 97 bytes/line).
FIELDS = [
    ("facility_code", 9),    # AS_FACILITY_CODE (the level ID)
    ("enumerator_id", 20),   # AS_ENUMERATOR_ID
    ("instrument", 4),       # AS_INSTRUMENT
    ("target_count", 4),     # AS_TARGET_COUNT
    ("ea_name", 50),         # AS_EA_NAME
    ("cluster", 10),         # AS_CLUSTER
]
RECORD_LEN = sum(n for _, n in FIELDS)
INSTRUMENTS = {"F1", "F3", "F4"}
COLUMNS = ["enumerator_id", "enumerator_name", "facility_code",
           "instrument", "target_count", "ea_name", "cluster"]


def fw(value, width: int) -> str:
    """Left-justify alpha into a fixed byte width (truncate / space-pad)."""
    s = "" if value is None else str(value)
    return s[:width].ljust(width)


def read_master(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as f:
            return [dict(r) for r in csv.DictReader(f)]
    # .xlsx via openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    if "Assignments" not in wb.sheetnames:
        raise SystemExit(f"'{path.name}' has no 'Assignments' sheet.")
    ws = wb["Assignments"]
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {name: header.index(name) for name in COLUMNS if name in header}
    missing = [c for c in COLUMNS if c not in idx and c != "enumerator_name"]
    if missing:
        raise SystemExit(f"'Assignments' is missing column(s): {missing}")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append({name: (r[idx[name]] if name in idx and idx[name] < len(r) else None)
                     for name in COLUMNS})
    return rows


def clean(rows: list[dict]) -> tuple[list[dict], list[str]]:
    """Validate + normalise; return (good rows, warnings)."""
    good, warns = [], []
    seen = set()
    for n, raw in enumerate(rows, start=2):
        eid = (str(raw.get("enumerator_id") or "")).strip()
        fac = (str(raw.get("facility_code") or "")).strip()
        inst = (str(raw.get("instrument") or "")).strip().upper()
        ea = (str(raw.get("ea_name") or "")).strip()
        cluster = (str(raw.get("cluster") or "")).strip()
        ename = (str(raw.get("enumerator_name") or "")).strip()
        tgt_raw = raw.get("target_count")
        # blank row or a footnote/instruction row (no EA, no instrument, no target) → ignore quietly
        if not fac and not inst and not str(tgt_raw or "").strip():
            continue
        where = f"row {n} ({eid or '?'} / {fac or '?'})"

        if not eid:
            warns.append(f"SKIP {where}: no enumerator_id."); continue
        if not (fac.isdigit() and len(fac) == 9):
            warns.append(f"SKIP {where}: facility_code must be 9 digits, got {fac!r}."); continue
        if inst not in INSTRUMENTS:
            warns.append(f"SKIP {where}: instrument must be F1/F3/F4, got {inst!r}."); continue
        try:
            tgt = int(float(tgt_raw))
        except (TypeError, ValueError):
            warns.append(f"SKIP {where}: target_count not a number ({tgt_raw!r})."); continue
        if tgt < 1:
            warns.append(f"SKIP {where}: target_count must be >= 1 (got {tgt})."); continue
        if tgt > 999:
            warns.append(f"{where}: target {tgt} > 999 won't fit a 3-digit sequence — capped at 999.")
            tgt = 999
        key = (eid, fac, inst)
        if key in seen:
            warns.append(f"{where}: duplicate (enumerator, EA, instrument) — included anyway.")
        seen.add(key)
        if ename and len(ename) > 0:
            pass
        if len(ea) > 50:
            warns.append(f"{where}: ea_name >50 chars — truncated in the .dat file.")
        good.append({"enumerator_id": eid, "enumerator_name": ename, "facility_code": fac,
                     "instrument": inst, "target_count": tgt, "ea_name": ea, "cluster": cluster})
    return good, warns


def write_dat_files(rows: list[dict], out: Path) -> dict[str, int]:
    by_enum: dict[str, list[dict]] = {}
    for r in rows:
        by_enum.setdefault(r["enumerator_id"], []).append(r)
    for eid, ers in sorted(by_enum.items()):
        lines = []
        for r in ers:
            line = "".join(fw(r[name], n) for name, n in FIELDS)
            assert len(line) == RECORD_LEN, (eid, len(line))
            lines.append(line)
        text = "\r\n".join(lines) + "\r\n"
        # newline="" disables text-mode translation so the explicit CRLF is kept
        # verbatim (otherwise Windows turns "\r\n" into "\r\r\n" and injects blank
        # records that break CSPro's fixed-width reader).
        with (out / f"AS_{eid}.dat").open("w", encoding="latin-1", errors="replace", newline="") as f:
            f.write(text)
    return {eid: len(v) for eid, v in by_enum.items()}


def case_keys(row: dict) -> list[str]:
    fac = row["facility_code"]
    return [f"{fac}{seq:03d}" for seq in range(1, row["target_count"] + 1)]


def write_case_keys_csv(rows: list[dict], out: Path) -> int:
    total = 0
    with (out / "case-keys.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["case_key", "enumerator_id", "enumerator_name", "instrument",
                    "facility_code", "ea_name", "cluster", "sequence"])
        for r in sorted(rows, key=lambda x: (x["enumerator_id"], x["facility_code"], x["instrument"])):
            for seq, key in enumerate(case_keys(r), start=1):
                w.writerow([key, r["enumerator_id"], r["enumerator_name"], r["instrument"],
                            r["facility_code"], r["ea_name"], r["cluster"], f"{seq:03d}"])
                total += 1
    return total


def write_html_sheets(rows: list[dict], out: Path) -> None:
    by_enum: dict[str, list[dict]] = {}
    for r in rows:
        by_enum.setdefault(r["enumerator_id"], []).append(r)

    css = """
    body{font-family:Arial,Helvetica,sans-serif;color:#1a1a1a;margin:0;}
    .sheet{padding:26px 30px;page-break-after:always;}
    .sheet:last-child{page-break-after:auto;}
    h1{color:#00532f;font-size:20px;margin:0 0 2px;}
    .meta{color:#555;font-size:12px;margin-bottom:14px;}
    .ea{border:1px solid #cfd8d3;border-radius:6px;margin:0 0 14px;overflow:hidden;}
    .ea-h{background:#e3f0ea;padding:7px 11px;font-size:13px;}
    .ea-h b{color:#00532f;}
    .keys{display:flex;flex-wrap:wrap;gap:6px 14px;padding:10px 12px;}
    .key{font-family:'Courier New',monospace;font-size:13px;border:1px solid #d7d7d7;
         border-radius:4px;padding:2px 7px;background:#fafafa;min-width:118px;}
    .key .seq{color:#888;}
    .note{font-size:11px;color:#777;margin-top:6px;}
    @media print{.sheet{padding:14mm;}}
    """
    parts = ["<!doctype html><html><head><meta charset='utf-8'>",
             "<title>CAPI assignment sheets</title>", f"<style>{css}</style></head><body>"]
    for eid, ers in sorted(by_enum.items()):
        name = next((e["enumerator_name"] for e in ers if e["enumerator_name"]), "")
        n_cases = sum(e["target_count"] for e in ers)
        parts.append("<div class='sheet'>")
        parts.append(f"<h1>Assignment — {escape(eid)}"
                     + (f" · {escape(name)}" if name else "") + "</h1>")
        parts.append(f"<div class='meta'>{len(ers)} EA(s) · {n_cases} case(s) total · "
                     "type each 12-digit key when you start a case · use only the keys listed here</div>")
        for e in ers:
            parts.append("<div class='ea'>")
            parts.append("<div class='ea-h'>"
                         f"<b>{escape(e['instrument'])}</b> · EA <b>{escape(e['facility_code'])}</b> · "
                         f"{escape(e['ea_name']) or '(no name)'} · cluster {escape(e['cluster']) or '—'} · "
                         f"target <b>{e['target_count']}</b></div>")
            parts.append("<div class='keys'>")
            for seq, key in enumerate(case_keys(e), start=1):
                parts.append(f"<span class='key'>{escape(key)} <span class='seq'>#{seq:03d}</span></span>")
            parts.append("</div></div>")
        parts.append("<div class='note'>Generated from assignment-master.xlsx · "
                     "Supervisor &amp; Enumerator hub. Confirm counts against the coverage report "
                     "before closing the site.</div>")
        parts.append("</div>")
    parts.append("</body></html>")
    (out / "assignment-sheets.html").write_text("".join(parts), encoding="utf-8")


def write_manifest(rows: list[dict], dat_counts: dict[str, int], total_keys: int,
                   warns: list[str], out: Path) -> None:
    by_enum: dict[str, list[dict]] = {}
    for r in rows:
        by_enum.setdefault(r["enumerator_id"], []).append(r)
    lines = ["CAPI assignment generation — manifest", "=" * 40, ""]
    lines.append(f"Enumerators: {len(by_enum)} · EA rows: {len(rows)} · total case keys: {total_keys}")
    lines.append("")
    for eid, ers in sorted(by_enum.items()):
        per_inst: dict[str, int] = {}
        for e in ers:
            per_inst[e["instrument"]] = per_inst.get(e["instrument"], 0) + e["target_count"]
        inst_str = ", ".join(f"{k}:{v}" for k, v in sorted(per_inst.items()))
        lines.append(f"  AS_{eid}.dat — {len(ers)} EA(s), {sum(e['target_count'] for e in ers)} cases ({inst_str})")
    lines += ["", "Files in this folder:",
              "  AS_<id>.dat          -> hub 'Assign Enumeration Area' serves these over Bluetooth.",
              "  assignment-sheets.html -> print; hand each enumerator their page (type the keys).",
              "  case-keys.csv        -> QA / CSWeb cross-check of every 12-digit key.",
              ""]
    if warns:
        lines.append(f"Warnings ({len(warns)}):")
        lines += [f"  - {w}" for w in warns]
    else:
        lines.append("No warnings — all rows valid.")
    (out / "manifest.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Generate hub .dat files + printable sheets from the assignment master.")
    ap.add_argument("--master", default=str(Path(__file__).resolve().with_name("assignment-master.xlsx")))
    ap.add_argument("--out", default=str(Path(__file__).resolve().with_name("out")))
    args = ap.parse_args(argv)

    master = Path(args.master)
    if not master.exists():
        raise SystemExit(f"master not found: {master}")
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    rows, warns = clean(read_master(master))
    if not rows:
        print("No valid assignment rows found.")
        for w in warns:
            print("  ! " + w)
        return 1

    dat_counts = write_dat_files(rows, out)
    total_keys = write_case_keys_csv(rows, out)
    write_html_sheets(rows, out)
    write_manifest(rows, dat_counts, total_keys, warns, out)

    print(f"OK — {len(dat_counts)} enumerator file(s), {len(rows)} EA row(s), {total_keys} case key(s)")
    print(f"   out: {out}")
    for eid, n in sorted(dat_counts.items()):
        print(f"   AS_{eid}.dat ({n} EA row(s))")
    if warns:
        print(f"   {len(warns)} warning(s) — see manifest.txt")
    return 0


if __name__ == "__main__":
    sys.exit(main())
