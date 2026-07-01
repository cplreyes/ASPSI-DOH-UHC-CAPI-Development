#!/usr/bin/env python3
"""
Build the CAPI assignment-master spreadsheet (`assignment-master.xlsx`).

This is the supervisor's prep sheet: one row per (enumerator x EA/facility),
holding exactly the fields the deployed Supervisor hub's ASSIGNMENT_DICT needs
(AS_FACILITY_CODE / AS_ENUMERATOR_ID / AS_INSTRUMENT / AS_TARGET_COUNT /
AS_EA_NAME / AS_CLUSTER). Fill it, then run `generate_assignments.py` to emit
the per-enumerator AS_<id>.dat hub files + printable case-key sheets.

Regenerate the blank template anytime:  python build_assignment_template.py
(Generator-over-hand-edit: edit THIS, not the .xlsx by hand, for structure.)
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().with_name("assignment-master.xlsx")

FONT = "Arial"
GREEN = "00532F"        # DOH-green header
GREEN_LT = "E3F0EA"
INPUT_BLUE = "0000FF"   # convention: hardcoded inputs in blue text
DERIVED_BG = "F2F2F2"
HDR = Font(name=FONT, bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols, fill=GREEN):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build():
    wb = Workbook()

    # ---- Sheet 1: README ---------------------------------------------------
    rd = wb.active
    rd.title = "README"
    rd.sheet_view.showGridLines = False
    lines = [
        ("CAPI Assignment Master — Supervisor prep sheet", "title"),
        ("ASPSI-DOH UHC Survey Year 2 · F1 / F3 / F4 · Supervisor & Enumerator hub", "sub"),
        ("Pre-loaded with the Los Baños pre-test assignments (RA 'Unique Question Number' list,", "sub"),
        ("municipality 040341). R6 roster names reused. Edit / add rows for later rounds.", "sub"),
        ("", "n"),
        ("What this is", "h"),
        ("One row per (enumerator x EA/facility). It carries exactly what the hub's assignment", "n"),
        ("record needs. Fill the Assignments sheet, then run generate_assignments.py to produce:", "n"),
        ("   •  AS_<enumerator_id>.dat  — the per-enumerator file the hub's 'Assign Enumeration", "n"),
        ("      Area' serves over Bluetooth (enumerator pulls it via 'Receive Assigned Data').", "n"),
        ("   •  assignment-sheets.html — printable case-key sheets (the reliable hand-out path).", "n"),
        ("   •  case-keys.csv          — every 12-digit case key, for QA / CSWeb cross-check.", "n"),
        ("", "n"),
        ("You assign AREAS + TARGETS, not individual respondents", "h"),
        ("The unit is the EA / facility, with a target case count — NOT a named respondent.", "n"),
        ("F1 = one facility head per facility · F3 = patients recruited on-site up to target ·", "n"),
        ("F4 = households up to target (members are rostered inside the interview). Respondent /", "n"),
        ("household pre-listing is descoped from hub v1, so nobody is assigned by name.", "n"),
        ("", "n"),
        ("The case key", "h"),
        ("12 digits = the 9-digit EA/facility code (real PSGC) + a 3-digit sequence.", "n"),
        ("When first/last_case_key are filled (as here, from the RA QN list) they are AUTHORITATIVE —", "n"),
        ("the generator emits those exact keys verbatim (the RA numbering isn't always 001-based:", "n"),
        ("facility-head keys end 000, and Brgy. Bayog runs 601..620). Leave first/last blank and the", "n"),
        ("generator falls back to code + 001..target. A wrong PSGC prefix is hard-rejected on the tablet.", "n"),
        ("", "n"),
        ("Columns to fill (Assignments sheet)", "h"),
        ("   supervisor_id    the enumerator's supervisor (fs-01 / fs-02). Review-only — shows the chain.", "n"),
        ("   enumerator_id    login of the enumerator (e.g. se-001). Names the .dat file + the row.", "n"),
        ("   enumerator_name  label only (for the printed sheet); not sent to the tablet.", "n"),
        ("   facility_code    9-digit EA/facility code, real PSGC (RRPPMMMFF).", "n"),
        ("   instrument       F1, F3, or F4.", "n"),
        ("   target_count     how many cases to complete at that EA.", "n"),
        ("   ea_name          EA / facility name (label).", "n"),
        ("   cluster          cluster label (free text).", "n"),
        ("   first/last key   the exact 12-digit QN range from the RA list. Authoritative — emitted verbatim.", "n"),
        ("   RA confirmed?    REVIEWER fills — 'OK' per verified row, or 'needs fix'.", "n"),
        ("   RA notes         REVIEWER fills — corrections (real facility code, target, name…).", "n"),
        ("", "n"),
        ("Two distribution paths (field-protocol choice)", "h"),
        ("   1.  Assignment sheet  — print assignment-sheets.html; enumerator TYPES the 12-digit", "n"),
        ("       keys. Needs no pairing; the PSGC gate protects key integrity. RELIABLE TODAY.", "n"),
        ("   2.  Hub Bluetooth     — 'Assign Enumeration Area' serves AS_<id>.dat; enumerator runs", "n"),
        ("       'Receive Assigned Data'. Nicer, no internet — but DEVICE-VERIFY PENDING (the", "n"),
        ("       syncfile-over-Bluetooth leg hasn't been run on the 2-tablet rig yet).", "n"),
        ("", "n"),
        ("Recommendation: use the printed sheet for the pretest; switch to Bluetooth once the", "n"),
        ("Assign/Receive leg is device-verified.", "n"),
    ]
    r = 1
    for text, kind in lines:
        cell = rd.cell(row=r, column=1, value=text)
        if kind == "title":
            cell.font = Font(name=FONT, bold=True, size=15, color=GREEN)
        elif kind == "sub":
            cell.font = Font(name=FONT, italic=True, size=10, color="555555")
        elif kind == "h":
            cell.font = Font(name=FONT, bold=True, size=11, color=GREEN)
        else:
            cell.font = Font(name=FONT, size=10, color="222222")
        r += 1
    rd.column_dimensions["A"].width = 95

    # ---- Sheet 2: Assignments ---------------------------------------------
    ws = wb.create_sheet("Assignments")
    headers = ["supervisor_id", "enumerator_id", "enumerator_name", "facility_code",
               "instrument", "target_count", "ea_name", "cluster",
               "first_case_key", "last_case_key", "RA confirmed?", "RA notes / corrections"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    # the two RA-review columns (K, L) get an amber header so it's obvious the reviewers fill them
    for c in (11, 12):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor="B45309")
    ws.freeze_panes = "A2"

    # Los Baños pre-test assignments — the RA "Unique Question Number for Pre-testing" QN list
    # mapped onto the R6 roster names (Carl: reuse R6 roster · F1 heads in scope). All in
    # municipality 040341. Four facilities each get ONE enumerator who does the F1 Facility Head
    # (key ...000) AND the 10 F3 patients (...001-010); Aly operates BOTH F4 barangays on her two
    # accounts — Bayog = se-003, Mayondon = se-006 (Carl 2026-07-01: Mayondon moved from Merlyne to
    # Aly but KEPT on the se-006 account, so the two-tablet structure stays). first/last_case_key are the RA's exact
    # QNs and are AUTHORITATIVE — the generator emits them verbatim (Bayog runs 601..620, heads
    # end 000, so the old code+001 derivation would be wrong). supervisor_id keeps the R6 teams
    # (Team A = fs-01 · Team B = fs-02); it's review-only — the hub Assignment record has no
    # supervisor field, so the generator ignores it.
    #              supervisor, enum,      name,      facility,    instr,target, ea_name,                              cluster,  first_case_key, last_case_key
    assignments = [
        ["fs-01", "se-001", "Pat",     "040341110", "F1", 1,  "Los Banos Doctors Hospital - Facility Head",     "040341", "040341110000", "040341110000"],
        ["fs-01", "se-001", "Pat",     "040341110", "F3", 10, "Los Banos Doctors Hospital - Patient",           "040341", "040341110001", "040341110010"],
        ["fs-01", "se-002", "Shan",    "040341120", "F1", 1,  "Laguna Provincial Hospital Bay - Facility Head", "040341", "040341120000", "040341120000"],
        ["fs-01", "se-002", "Shan",    "040341120", "F3", 10, "Laguna Provincial Hospital Bay - Patient",       "040341", "040341120001", "040341120010"],
        ["fs-01", "se-004", "Marriz",  "040341130", "F1", 1,  "Los Banos RHU I - Facility Head",                "040341", "040341130000", "040341130000"],
        ["fs-01", "se-004", "Marriz",  "040341130", "F3", 10, "Los Banos RHU I - Patient",                      "040341", "040341130001", "040341130010"],
        ["fs-01", "se-003", "Aly",     "040341100", "F4", 20, "Brgy. Bayog - Household",                        "040341", "040341100601", "040341100620"],
        ["fs-02", "se-005", "Aidan",   "040341140", "F1", 1,  "St. Jude Hospital - Facility Head",              "040341", "040341140000", "040341140000"],
        ["fs-02", "se-005", "Aidan",   "040341140", "F3", 10, "St. Jude Hospital - Patient",                    "040341", "040341140001", "040341140010"],
        ["fs-02", "se-006", "Aly",     "040341101", "F4", 20, "Brgy. Mayondon - Household",                     "040341", "040341101001", "040341101020"],
    ]
    for i, row in enumerate(assignments, start=2):
        ws.append(row)
        # first/last_case_key (cols I/J) are literal 12-digit QNs — keep as text so Excel
        # doesn't strip the leading zero or coerce to a number.
        ws.cell(row=i, column=9).number_format = "@"
        ws.cell(row=i, column=10).number_format = "@"

    # column widths (A supervisor_id … L RA notes)
    widths = [14, 14, 16, 14, 11, 12, 42, 10, 15, 15, 14, 36]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w

    # styling: input cols (A-H) blue, derived cols (I-J) grey, RA cols (K-L) amber
    input_cols = [1, 2, 3, 4, 5, 6, 7, 8]
    center_cols = {1, 2, 4, 5, 6, 8}  # ids/codes/instr/target/cluster centered; names left
    last = 1 + len(assignments)
    for rr in range(2, last + 1):
        for c in input_cols:
            cell = ws.cell(row=rr, column=c)
            cell.font = Font(name=FONT, size=10, color=INPUT_BLUE)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if c in center_cols else "left",
                                       vertical="center")
        for c in (9, 10):
            cell = ws.cell(row=rr, column=c)
            cell.font = Font(name=FONT, size=10, color="000000")
            cell.fill = PatternFill("solid", fgColor=DERIVED_BG)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in (11, 12):  # RA review columns — light amber = reviewer fills
            cell = ws.cell(row=rr, column=c)
            cell.font = Font(name=FONT, size=10, color="000000")
            cell.fill = PatternFill("solid", fgColor="FFF3D6")
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if c == 11 else "left", vertical="center")
        ws.cell(row=rr, column=6).number_format = "0"

    # instrument dropdown (col E)
    dv = DataValidation(type="list", formula1='"F1,F3,F4"', allow_blank=False)
    dv.error = "Choose F1, F3, or F4"
    dv.prompt = "F1 = Facility Head · F3 = Patient · F4 = Household"
    ws.add_data_validation(dv)
    dv.add("E2:E200")
    # target whole-number > 0 (col F)
    dvt = DataValidation(type="whole", operator="greaterThan", formula1="0", allow_blank=False)
    dvt.error = "Target must be a whole number greater than 0"
    ws.add_data_validation(dvt)
    dvt.add("F2:F200")
    # RA confirmed? dropdown
    dvc = DataValidation(type="list", formula1='"OK,needs fix"', allow_blank=True)
    dvc.prompt = "Reviewer: 'OK' once verified, or 'needs fix' (put the correction in RA notes)"
    ws.add_data_validation(dvc)
    dvc.add("K2:K200")

    # note row
    note_r = last + 2
    nc = ws.cell(row=note_r, column=1,
                 value="↑ Los Baños pre-test (RA QN list, municipality 040341; R6 roster names). "
                       "4 facilities = F1 head + F3 patients · 2 barangays = F4. "
                       "Blue = you fill · grey = QN range from the RA list (verbatim). Edit / add rows for later rounds.")
    nc.font = Font(name=FONT, italic=True, size=9, color="555555")
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=12)

    # ---- Sheet 3: Enumerators (reference roster) ---------------------------
    en = wb.create_sheet("Enumerators")
    en.append(["enumerator_id", "name", "role", "team"])
    style_header(en, 1, 4)
    roster = [
        ["fs-01", "Aidan", "Supervisor QA", "Team A"],
        ["se-001", "Pat", "Field Sync", "Team A"],
        ["se-002", "Shan", "Field Sync", "Team A"],
        ["se-003", "Aly", "Field Sync", "Team A"],
        ["se-004", "Marriz", "Field Sync", "Team A"],
        ["fs-02", "Ms. Marriz", "Supervisor QA", "Team B"],
        ["se-005", "Aidan", "Field Sync", "Team B"],
        ["se-006", "Merlyne", "Field Sync", "Team B"],
    ]
    for row in roster:
        en.append(row)
    for idx, w in enumerate([14, 16, 16, 10], start=1):
        en.column_dimensions[chr(64 + idx)].width = w
    en.freeze_panes = "A2"
    for rr in range(2, 2 + len(roster)):
        for c in range(1, 5):
            en.cell(row=rr, column=c).font = Font(name=FONT, size=10)
            en.cell(row=rr, column=c).border = BORDER
    note = ("R6 roster (names only, no passwords). Use these IDs in the Assignments sheet. "
            "Dual-role on opposite teams: Aidan = fs-01 + se-005 · Marriz = fs-02 + se-004.")
    en.cell(row=2 + len(roster) + 1, column=1, value=note
            ).font = Font(name=FONT, italic=True, size=9, color="555555")

    # README is internal guidance, not the reviewers' concern — hide it and make
    # the workbook open on the Assignments sheet.
    rd.sheet_state = "hidden"
    wb.active = wb.sheetnames.index("Assignments")

    # Force Excel/LibreOffice to compute the derived-key formulas on open
    # (no pre-recalc step needed; the AF_UNIX recalc helper is Linux-only).
    wb.calculation.fullCalcOnLoad = True

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    build()
