#!/usr/bin/env python3
"""Export a CSPro .dcf (JSON) to an Excel workbook using the same two-sheet
layout CSPro Designer produces: "Dictionary Names and Labels" + "Value Sets".

Matches the format of raw/Sample-Data/Facility Head Survey Data Dictionary
and Value Sets.xlsx so reviewers open it without surprise.

Usage:
    python export_dcf_to_xlsx.py <file.dcf> [-o <out.xlsx>]
    python export_dcf_to_xlsx.py --all
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent

BATCH_TARGETS = [
    HERE / "F1" / "FacilityHeadSurvey.dcf",
    HERE / "F3" / "PatientSurvey.dcf",
    HERE / "F4" / "HouseholdSurvey.dcf",
]

NAMES_SHEET = "Dictionary Names and Labels"
VALUES_SHEET = "Value Sets"

NAMES_WIDTHS = [30, 30, 40, 40, 55, 80]
VALUES_WIDTHS = [45, 80, 60, 10]


def _label(obj: dict) -> str:
    labels = obj.get("labels") or []
    if labels and isinstance(labels[0], dict):
        return labels[0].get("text", "") or ""
    return ""


def _pair_value(pair: dict) -> str:
    if "value" in pair:
        return str(pair["value"])
    if "from" in pair and "to" in pair:
        return f"{pair['from']}-{pair['to']}"
    if "from" in pair:
        return str(pair["from"])
    return ""


def _ids_tag(level_index: int) -> str:
    return f"_IDS{level_index}"


def write_names_sheet(ws, dct: dict) -> None:
    ws.title = NAMES_SHEET

    header_font = Font(bold=True)

    for i, w in enumerate(NAMES_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for li, level in enumerate(dct.get("levels") or []):
        # Level header row: A=name, B=label
        row = [level.get("name", ""), _label(level), None, None, None, None]
        ws.append(row)
        ws.cell(row=ws.max_row, column=1).font = header_font

        # Id items pseudo-record
        ids = level.get("ids") or {}
        id_items = ids.get("items") or []
        if id_items:
            ws.append([None, None, _ids_tag(li), "(Id Items)", None, None])
            for it in id_items:
                ws.append([None, None, None, None, it.get("name", ""), _label(it)])

        # Regular records
        for rec in level.get("records") or []:
            ws.append([None, None, rec.get("name", ""), _label(rec), None, None])
            for it in rec.get("items") or []:
                ws.append([None, None, None, None, it.get("name", ""), _label(it)])
                for sub in it.get("subitems") or []:
                    ws.append([None, None, None, None, sub.get("name", ""), _label(sub)])

    # Left-align + wrap labels for readability
    wrap = Alignment(wrap_text=False, vertical="top")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap


def iter_valuesets(dct: dict):
    """Yield (valueset_name, valueset_label, [(value_label, code), ...]) in DCF order."""
    for level in dct.get("levels") or []:
        for rec in level.get("records") or []:
            for it in rec.get("items") or []:
                for vs in it.get("valueSets") or []:
                    pairs = []
                    for v in vs.get("values") or []:
                        lbl = _label(v)
                        codes = [_pair_value(p) for p in (v.get("pairs") or [])]
                        code_str = ", ".join(c for c in codes if c)
                        pairs.append((lbl, code_str))
                    yield vs.get("name", ""), _label(vs), pairs


def write_values_sheet(ws, dct: dict) -> None:
    ws.title = VALUES_SHEET

    header_font = Font(bold=True)

    for i, w in enumerate(VALUES_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for name, label, pairs in iter_valuesets(dct):
        ws.append([name, label, None, None])
        ws.cell(row=ws.max_row, column=1).font = header_font
        for vlbl, code in pairs:
            ws.append([None, None, vlbl, code])
        ws.append([None, None, None, None])  # blank row separator

    align = Alignment(vertical="top")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align


def export(dcf_path: Path, out_path: Path | None = None) -> Path:
    with dcf_path.open(encoding="utf-8") as fh:
        dct: dict[str, Any] = json.load(fh)

    wb = Workbook()
    write_names_sheet(wb.active, dct)
    write_values_sheet(wb.create_sheet(), dct)

    if out_path is None:
        out_path = dcf_path.with_name(f"{dcf_path.stem} - Data Dictionary and Value Sets.xlsx")

    wb.save(out_path)
    return out_path


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("dcf", nargs="?", type=Path, help="Path to .dcf file")
    ap.add_argument("-o", "--output", type=Path, help="Output .xlsx path")
    ap.add_argument("--all", action="store_true", help="Export F1/F3/F4 DCFs next to each source")
    args = ap.parse_args()

    if args.all:
        for target in BATCH_TARGETS:
            if not target.exists():
                print(f"SKIP (missing): {target}")
                continue
            out = export(target)
            print(f"OK: {target.name} -> {out.name}")
        return

    if args.dcf is None:
        ap.error("provide a .dcf path or use --all")

    out = export(args.dcf, args.output)
    print(f"OK: {args.dcf.name} -> {out}")


if __name__ == "__main__":
    main()
