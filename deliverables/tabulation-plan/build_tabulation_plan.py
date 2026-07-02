#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_tabulation_plan.py — UHC Survey Year 2 tabulation-plan generator (Epic 11).

SOURCE OF TRUTH:
  raw/Email-Ingest-2026-06-05/Attachments/2_DOH UHC Yr2_PSA SSRCS Form 1_signed.pdf
  §II item 9 "List of tables and other outputs to be generated" (PDF pages 3-9).

Descriptions below are transcribed VERBATIM from the signed form (including the
printed quirks: 2.49a/b duplicate 2.48's description, 3.24/3.25 identical,
4.4/4.5 identical, 1.21 and 4.9 "(continued)", no plain Table 1.17).

GENERATOR-FIRST RULE: edit THIS file and re-run it. Never hand-edit the
emitted tabulation-plan.xlsx / tabulation-plan.csv.

Run:  PYTHONIOENCODING=utf-8 python build_tabulation_plan.py
Emits: tabulation-plan.xlsx, tabulation-plan.csv (next to this script).
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent

ANNEX_INSTRUMENT = {1: "F1", 2: "F3", 3: "F4", 4: "F2"}
ANNEX_WEIGHT = {
    1: "facility wf",
    2: "patient wf",
    3: "household wf",
    4: "HCW wf",
}
ANNEX_TITLE = {
    1: "Annex 1. Facility-level indicators (F1)",
    2: "Annex 2. Patient-level indicators (F3)",
    3: "Annex 3. Household-level indicators (F4)",
    4: "Annex 4. Healthcare worker-level indicators (F2)",
}

DOH_GREEN = "00532F"

STATA12_CONSTRAINT = (
    "Target: Stata 12 SE (portable, C:\\Users\\analy\\Downloads\\Stata 12\\StataSE.exe; "
    "new Stata under procurement). Tabulation programs must use Stata 12 syntax: "
    "svyset + svy: tab / svy: proportion / svy: mean + tabout (user-written) for export. "
    "NOT available: putexcel (13+), collect/etable/new table (17+), import excel is "
    "available (12) but prefer the ETL's CSV/.dta. The harmonization ETL must write "
    ".dta format <=115 (pandas to_stata(version=114)); str244 max; 32-char variable "
    "names. All tables are WEIGHTED (SSRCS Annex A): base weight x stage weights x "
    "non-response adjustment = final weight wf; facility tables break down by facility "
    "type (911 RHU / 244 gov hospital / 366 private hospital of n=1,521), patient/HH "
    "tables by province."
)


def T(no, description, stat_type, universe, breakdown, notes=""):
    """Build one full table-definition dict; annex/instrument/weight derive from `no`."""
    annex = int(no.split(".")[0])
    return {
        "no": no,
        "annex": annex,
        "instrument": ANNEX_INSTRUMENT[annex],
        "description": description,
        "stat_type": stat_type,
        "universe": universe,
        "breakdown": breakdown,
        "weight": ANNEX_WEIGHT[annex],
        "source_variables": "",
        "mapping_status": "unmapped",
        "notes": notes,
    }


FH = "all facility heads (n=1,521)"
HCW = "HCWs in selected facilities"

TABLES = [
    # ------------------------------------------------------------------
    # Annex 1. Facility-level indicators (F1) — Tables 1.1-1.32 + 1.8a/b/c + 1.17a/b
    # ------------------------------------------------------------------
    T("1.1", "Percentage distribution of facility head respondents’ who have ever heard of the process to apply for PhilHealth YAKAP/Konsulta accreditation, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.2", "Percentage distribution of facility head respondents’ reported support from LGU to implement UHC reforms, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.3", "Percentage distribution of facility head respondents reported type of support from LGU to implement UHC reforms, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.4", "Percentage distribution of facility head respondents perceived satisfaction with the support received from LGU to implement UHC reforms, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.5", "Percentage distribution of facility head respondents reported reasons for dissatisfaction of support from LGU to implement UHC reforms",
      "percent distribution", "facility heads dissatisfied with LGU support", "none stated"),
    T("1.6", "Percentage distribution of facility head respondents who have ever heard of the process to apply for DOH licensing, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.7", "Percentage distribution of facility head respondents reported that the facility is currently PhilHealth YAKAP/Konsulta accredited, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.8", "Percentage distribution of facility head respondents identified top barriers in facilities' accreditation to PhilHealth YAKAP/Konsulta, for the most recent application, by type of facility (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.8a", "Percentage distribution of facility head respondents of rural health unit reported reasons to barriers in facilities' accreditation to PhilHealth YAKAP/Konsulta, for the most recent application (n=911)",
      "percent distribution", "facility heads of rural health units (n=911)", "none stated"),
    T("1.8b", "Percentage distribution of facility head respondents of government hospitals reported reasons to barriers in facilities' accreditation to PhilHealth YAKAP/Konsulta, for the most recent application (n=244)",
      "percent distribution", "facility heads of government hospitals (n=244)", "none stated"),
    T("1.8c", "Percentage distribution of facility head respondents of private hospitals reported reasons to barriers in facilities' accreditation to PhilHealth YAKAP/Konsulta, for the most recent application (n=366)",
      "percent distribution", "facility heads of private hospitals (n=366)", "none stated"),
    T("1.9", "Percentage distribution of the average number of days to PhilHealth YAKAP/Konsulta accreditation for the facility's most recent application, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.10", "Percentage distribution of basic services outsourced by YAKAP/Konsulta provider or primary care providers, over the past 6 months, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.11", "Percentage distribution of top methods of outbound referral, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.12", "Percentage distribution of top methods of inbound referral of hospitals, by facility type (n=1,521)",
      "percent distribution", "hospitals", "by facility type"),
    T("1.13", "Percent distribution of primary care providers who have a network of specialist providers or referral health facilities to refer patients to, by facility type (n=1,521)",
      "percent distribution", "primary care providers", "by facility type"),
    T("1.14", "Percent distribution of the facility head respondents’ satisfaction rate for the referral system over the past 6 months, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.15", "Percent distribution of top challenges for facilities in the referral system, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.16", "Percent distribution of facilities that are DOH licensed, at the point of survey, by facility type (n=1,521)",
      "percent distribution", "all facilities (n=1,521)", "by facility type"),
    T("1.17a", "Percent distribution of top barriers in DOH licensing for rural/city health unit, for the most recent application (n=911)",
      "percent distribution", "rural/city health units (n=911)", "none stated",
      "no parent Table 1.17 printed; only 1.17a/1.17b"),
    T("1.17b", "Percent distribution of top barriers in DOH licensing for hospitals, for the most recent application",
      "percent distribution", "hospitals", "none stated",
      "no n= printed (cf. 1.17a n=911)"),
    T("1.18", "Percentage distribution of the average number of days to DOH licensing for the facility's most recent application, by facility type (n=1,521)",
      "percent distribution", FH, "by facility type"),
    T("1.19", "Percentage distribution of facilities who provide NBB to 100% of eligible patients, over the past 6 months, by facility type",
      "percent distribution", "facilities", "by facility type"),
    T("1.20", "Percentage distribution of primary care facilities experiencing stock-outs of tracer essential medicines (e.g., antihypertensives, antibiotics) for more than 7 days in the past quarter",
      "percent distribution", "primary care facilities", "none stated"),
    T("1.21", "Percentage distribution of primary care facilities experiencing stock-outs of tracer essential medicines (e.g., antihypertensives, antibiotics) for more than 7 days in the past quarter (continued)",
      "percent distribution", "primary care facilities", "none stated",
      "printed as (continued) of Table 1.20"),
    T("1.22", "Percentage distribution of patients successfully referred to a higher-level facility within the network (HCPN) who received care",
      "percent distribution", "patients referred to higher-level facility within HCPN", "none stated"),
    T("1.23", "Percentage distribution of facilities submitting routine health and financial data to DOH Information System and/or PhilHealth Dashboard",
      "percent distribution", "facilities", "none stated"),
    T("1.24", "Percentage distribution on the frequency of health and financial data submissions to DOH Information System and/or PhilHealth Dashboard",
      "percent distribution", "facilities", "none stated"),
    T("1.25", "Percentage distribution of the submitted health and financial data reports used for decision-making",
      "percent distribution", "facilities", "none stated"),
    T("1.26", "Percentage distribution of facility head respondents who report clearly defined roles and autonomy vis-à-vis the LGU/Provincial Health Office",
      "percent distribution", "facility heads", "none stated"),
    T("1.27", "Percentage distribution of facility head respondents who reported delays in receiving capitation tranches",
      "percent distribution", "facility heads", "none stated"),
    T("1.28", "Percentage distribution on the average time interval between tranches releases to the facility",
      "percent distribution", "facilities", "none stated"),
    T("1.29", "Percentage distribution of facility heads by role in the health facility",
      "percent distribution", "facility heads", "by role in the health facility"),
    T("1.30", "Percentage distribution of facility heads by age group",
      "percent distribution", "facility heads", "by age group"),
    T("1.31", "Percentage distribution of facility heads by sex at birth",
      "percent distribution", "facility heads", "by sex at birth"),
    T("1.32", "Percentage distribution of health facilities by service capacity level",
      "percent distribution", "health facilities", "by service capacity level"),

    # ------------------------------------------------------------------
    # Annex 2. Patient-level indicators (F3) — Tables 2.1-2.65 (many a/b splits)
    # ------------------------------------------------------------------
    T("2.1", "Percent of inpatient and outpatient respondents who have heard about UHC, by province",
      "percent", "inpatients+outpatients", "by province × in/outpatient"),
    T("2.2a", "Percent distribution of source of information about UHC of inpatient respondents, by province",
      "percent distribution", "inpatients", "by province"),
    T("2.2b", "Percent distribution of source of information about UHC of outpatient respondents, by province",
      "percent distribution", "outpatients", "by province"),
    T("2.3", "Percent distribution of inpatient and outpatient respondents who identify as registered to PhilHealth, by province",
      "percent distribution", "inpatients+outpatients", "by province × in/outpatient"),
    T("2.4", "Percentage distribution of unregistered inpatient and outpatient who know where to seek assistance for the registration process, by province",
      "percent distribution", "unregistered inpatients+outpatients", "by province × in/outpatient"),
    T("2.5a", "Percentage distribution of inpatient respondents’ source of information on PhilHealth registration, by province",
      "percent distribution", "inpatients", "by province"),
    T("2.5b", "Percentage distribution of outpatient respondents’ source of information on PhilHealth registration, by province",
      "percent distribution", "outpatients", "by province"),
    T("2.6a", "Percentage distribution of inpatient respondents who can correctly list PhilHealth benefits under UHC reforms, by province",
      "percent distribution", "inpatients", "by province"),
    T("2.6b", "Percentage distribution of outpatient respondents who can correctly list PhilHealth benefits under UHC reforms, by province",
      "percent distribution", "outpatients", "by province"),
    T("2.7", "Percentage distribution of inpatient and outpatient respondents who have heard about YAKAP/Konsulta provider assignment under UHC, by province",
      "percent distribution", "inpatients+outpatients", "by province × in/outpatient"),
    T("2.8a", "Percentage distribution of inpatient respondents who can identify the benefits under the PhilHealth YAKAP/Konsulta package",
      "percent distribution", "inpatients", "none stated"),
    T("2.8b", "Percentage distribution of outpatient respondents who can identify the benefits under the PhilHealth YAKAP/Konsulta package",
      "percent distribution", "outpatients", "none stated"),
    T("2.9a", "Percentage distribution of inpatient respondents who state that they would visit the YAKAP/Konsulta provider or primary care provider as their first point of contact, by province",
      "percent distribution", "inpatients", "by province"),
    T("2.9b", "Percentage distribution of outpatient respondents who state that they would visit the YAKAP/Konsulta provider or primary care provider as their first point of contact, by province",
      "percent distribution", "outpatients", "by province"),
    T("2.10a", "Percentage distribution of inpatient respondents’ top challenges in registration to PhilHealth for those who have undergone the registration process (including those who did not register successfully) in the past year, by province",
      "percent distribution", "inpatients who have undergone the registration process", "by province"),
    T("2.10b", "Percentage distribution of outpatient respondents’ top challenges in registration to PhilHealth for those who have undergone the registration process (including those who did not register successfully) in the past year, by province",
      "percent distribution", "outpatients who have undergone the registration process", "by province"),
    T("2.11a", "Percent distribution of inpatient respondents by type of PhilHealth member who self-report to have struggled to pay premiums in the past year, by province",
      "percent distribution", "inpatient PhilHealth members", "by PhilHealth member type × province"),
    T("2.11b", "Percent distribution of outpatient respondents by type of PhilHealth member who self-report to have struggled to pay premiums in the past year, by province",
      "percent distribution", "outpatient PhilHealth members", "by PhilHealth member type × province"),
    T("2.12", "Percent distribution of inpatient and outpatient respondents that have registered to a YAKAP/Konsulta provider, by province",
      "percent distribution", "inpatients+outpatients", "by province × in/outpatient"),
    T("2.13", "Percent distribution of inpatient and outpatient respondents who have consulted their assigned YAKAP/Konsulta provider since registering, by province",
      "percent distribution", "inpatients+outpatients registered to a YAKAP/Konsulta provider", "by province × in/outpatient"),
    T("2.14", "Percent distribution of inpatient and outpatient respondents in HCPN that were referred from YAKAP/Konsulta provider or primary care provider, for most recent interaction, by province",
      "percent distribution", "inpatients+outpatients in HCPN", "by province × in/outpatient"),
    T("2.15", "Percent distribution of inpatient and outpatient respondents that received assistance in making the appointment for that visit, by province",
      "percent distribution", "inpatients+outpatients", "by province × in/outpatient"),
    T("2.16", "Percent distribution of inpatient and outpatient respondents that received assistance by discussing with them the different places they could visit to address their health problem, by province",
      "percent distribution", "inpatients+outpatients", "by province × in/outpatient"),
    T("2.17a", "Percentage distribution of inpatient respondents’ overall satisfaction rate with the referral system process for most recent interaction at the time of survey, by province",
      "percent distribution", "inpatients", "by province"),
    T("2.17b", "Percentage distribution of outpatient respondents’ overall satisfaction rate with the referral system process for most recent interaction at the time of survey, by province",
      "percent distribution", "outpatients", "by province"),
    T("2.18", "Percent distribution of inpatient and outpatient respondents who can correctly define a generic drug, by province",
      "percent distribution", "inpatients+outpatients", "by province × in/outpatient"),
    T("2.19", "Percentage distribution of inpatient and outpatient respondents who chose a generic drug when buying medication, for last interaction, by province",
      "percent distribution", "inpatients+outpatients", "by province × in/outpatient"),
    T("2.20a", "Percentage distribution of inpatient respondents’ reasons to purchase generic options (prescribed by doctor, cheaper, etc.) in the past 6 months, by province",
      "percent distribution", "inpatients", "by province"),
    T("2.20b", "Percentage distribution of outpatient respondents’ reasons to purchase generic options (prescribed by doctor, cheaper, etc.) in the past 6 months, by province",
      "percent distribution", "outpatients", "by province"),
    T("2.21a", "Percentage distribution of inpatient respondents’ reasons to purchase branded options (prescribed by doctor, cheaper, etc.) in the past 6 months, by province",
      "percent distribution", "inpatients", "by province"),
    T("2.21b", "Percentage distribution of outpatient respondents’ reasons to purchase branded options (prescribed by doctor, cheaper, etc.) in the past 6 months, by province",
      "percent distribution", "outpatients", "by province"),
    T("2.22a", "Percent distribution of inpatient respondents that accessed medications for free or subsidized, in the last 6 months, by province",
      "percent distribution", "inpatients", "by province"),
    T("2.22b", "Percent distribution of outpatient respondents that accessed medications for free or subsidized, in the last 6 months, by province",
      "percent distribution", "outpatients", "by province"),
    T("2.23a", "Percent distribution of inpatient respondents’ top sources of medicine subsidization (i.e., LGU, employer, etc.) in the past 6 months",
      "percent distribution", "inpatients", "none stated"),
    T("2.23b", "Percent distribution of outpatient respondents’ top sources of medicine subsidization (i.e., LGU, employer, etc.) in the past 6 months",
      "percent distribution", "outpatients", "none stated"),
    T("2.24a", "Percentage distribution of inpatient respondents’ that have accessed to alternate financial assistance program/s to help cover medical costs, for most recent interaction, out of all patients",
      "percent distribution", "all inpatients", "none stated"),
    T("2.24b", "Percentage distribution of outpatient respondents’ that have accessed to alternate financial assistance program/s to help cover medical costs, for most recent interaction, out of all patients",
      "percent distribution", "all outpatients", "none stated"),
    T("2.25a", "Percentage distribution of inpatient respondents’ that have accessed to alternate financial assistance program/s to help cover medical costs, for most recent interaction, out of billed patients",
      "percent distribution", "billed inpatients", "none stated"),
    T("2.25b", "Percentage distribution of outpatient respondents’ that have accessed to alternate financial assistance program/s to help cover medical costs, for most recent interaction, out of billed patients",
      "percent distribution", "billed outpatients", "none stated"),
    T("2.26", "Percentage distribution of inpatient and outpatient respondents’ NBB-eligible who received NBB for most recent interaction",
      "percent distribution", "NBB-eligible inpatients+outpatients", "none stated"),
    T("2.27", "Percentage distribution of inpatient and outpatient respondents’ NBB-eligible who received ZBB for most recent interaction",
      "percent distribution", "NBB-eligible inpatients+outpatients", "none stated"),
    T("2.28a", "Summary of alternate financial assistance program/s covering the inpatients’ medical costs, in their most recent interaction",
      "summary", "inpatients", "none stated"),
    T("2.28b", "Summary of alternate financial assistance program/s covering the outpatients’ medical costs, in their most recent interaction",
      "summary", "outpatients", "none stated"),
    T("2.29", "Catastrophic health expenditure at the 25% and 40% thresholds across different subgroups among patients within UHC-IS and Non-UHC-IS",
      "percent", "inpatients+outpatients (UHC-IS and Non-UHC-IS)", "by subgroup × UHC-IS/Non-UHC-IS"),
    T("2.30", "Percentage distribution of inpatient and outpatient respondents that use loan/mortgage to pay for medical costs, for most recent interaction",
      "percent distribution", "inpatients+outpatients", "none stated"),
    T("2.31", "Percentage distribution of inpatient and outpatient respondents that use sale of assets to pay for medical costs, for most recent interaction",
      "percent distribution", "inpatients+outpatients", "none stated"),
    T("2.32", "Percentage distribution of inpatient and outpatient respondents that experience incidence of reduced care due to financial reasons among the population, for most recent interaction",
      "percent distribution", "inpatients+outpatients", "none stated"),
    T("2.33a", "Health care expenditures of inpatient respondents by type of facility, for the most recent interaction, by province",
      "summary", "inpatients", "by facility type × province"),
    T("2.33b", "Health care expenditures of outpatient respondents by type of facility, for the most recent interaction, by province",
      "summary", "outpatients", "by facility type × province"),
    T("2.34a", "Average time (in minutes) of inpatient respondents to access nearest health facility for the latest visit, by province",
      "average", "inpatients", "by province"),
    T("2.34b", "Average time (in minutes) of outpatient respondents to access nearest health facility for the latest visit, by province",
      "average", "outpatients", "by province"),
    T("2.35a", "Average costs spent of inpatient respondents on going to the nearest health facility for the latest visit, by level of care",
      "average", "inpatients", "by level of care"),
    T("2.35b", "Average costs spent of outpatient respondents on going to the nearest health facility for the latest visit, by level of care",
      "average", "outpatients", "by level of care"),
    T("2.36a", "Percentage distribution of inpatient respondents that walk as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.36b", "Percentage distribution of outpatient respondents that walk as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.37a", "Percentage distribution of inpatient respondents that ride a bike as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.37b", "Percentage distribution of outpatient respondents that ride a bike as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.38a", "Percentage distribution of inpatient respondents that ride a public bus as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.38b", "Percentage distribution of outpatient respondents that ride a public bus as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.39a", "Percentage distribution of inpatient respondents that ride a jeepney as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.39b", "Percentage distribution of outpatient respondents that ride a jeepney as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.40a", "Percentage distribution of inpatient respondents that ride a tricycle as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.40b", "Percentage distribution of outpatient respondents that ride a tricycle as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.41a", "Percentage distribution of inpatient respondents that ride a car as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.41b", "Percentage distribution of outpatient respondents that ride a car as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.42a", "Percentage distribution of inpatient respondents that ride a motorcycle as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.42b", "Percentage distribution of outpatient respondents that ride a motorcycle as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.43a", "Percentage distribution of inpatient respondents that ride a boat as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.43b", "Percentage distribution of outpatient respondents that ride a boat as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.44a", "Percentage distribution of inpatient respondents that ride a taxi as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.44b", "Percentage distribution of outpatient respondents that ride a taxi as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.45a", "Percentage distribution of inpatient respondents that ride a pedicab as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.45b", "Percentage distribution of outpatient respondents that ride a pedicab as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.46a", "Percentage distribution of inpatient respondents that ride an E-bike as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "inpatients", "none stated"),
    T("2.46b", "Percentage distribution of outpatient respondents that ride an E-bike as mode of transportation used on going to the nearest health facility for the latest visit",
      "percent distribution", "outpatients", "none stated"),
    T("2.47a", "Responsiveness score summary for inpatient respondents",
      "score", "inpatients", "none stated"),
    T("2.47b", "Responsiveness score summary for outpatient respondents",
      "score", "outpatients", "none stated"),
    T("2.48a", "Respect for Persons Domains: Dignity, Autonomy, Confidentiality for inpatient respondents",
      "score", "inpatients", "none stated"),
    T("2.48b", "Respect for Persons Domains: Dignity, Autonomy, Confidentiality for outpatient respondents",
      "score", "outpatients", "none stated"),
    T("2.49a", "Respect for Persons Domains: Dignity, Autonomy, Confidentiality for inpatient respondents",
      "score", "inpatients", "none stated",
      "duplicate description as printed (same as Table 2.48a)"),
    T("2.49b", "Respect for Persons Domains: Dignity, Autonomy, Confidentiality for outpatient respondents",
      "score", "outpatients", "none stated",
      "duplicate description as printed (same as Table 2.48b)"),
    T("2.50a", "Overall inpatient experience score across levels of facility, socioeconomic status, place of residence, and sex, for their most recent encounter",
      "score", "inpatients", "by facility level × SES × residence × sex"),
    T("2.50b", "Overall outpatient experience score across levels of facility, socioeconomic status, place of residence, and sex, for their most recent encounter",
      "score", "outpatients", "by facility level × SES × residence × sex"),
    T("2.51", "Average out-of-pocket payments on rooms of inpatient and outpatient respondents",
      "average", "inpatients+outpatients", "none stated"),
    T("2.52", "Average out-of-pocket payments on consultation of inpatient and outpatient respondents",
      "average", "inpatients+outpatients", "none stated"),
    T("2.53", "Average out-of-pocket payments on laboratory tests of inpatient and outpatient respondents",
      "average", "inpatients+outpatients", "none stated"),
    T("2.54", "Average out-of-pocket payments on prescribed medicines of inpatient and outpatient respondents",
      "average", "inpatients+outpatients", "none stated"),
    T("2.55", "Percentage distribution of inpatient and outpatient respondents who have heard about ZBB",
      "percent distribution", "inpatients+outpatients", "none stated"),
    T("2.56a", "Percentage distribution of inpatient respondents’ top information sources about ZBB",
      "percent distribution", "inpatients", "none stated"),
    T("2.56b", "Percentage distribution of outpatient respondents’ top information sources about ZBB",
      "percent distribution", "outpatients", "none stated"),
    T("2.57a", "Percentage distribution of inpatient respondents who can correctly identify the benefits under ZBB",
      "percent distribution", "inpatients", "none stated"),
    T("2.57b", "Percentage distribution of outpatient respondents who can correctly identify the benefits under ZBB",
      "percent distribution", "outpatients", "none stated"),
    T("2.58", "Percentage distribution of inpatient and outpatient respondents who have heard about BUCAS",
      "percent distribution", "inpatients+outpatients", "none stated"),
    T("2.59a", "Percentage distribution of inpatient respondents’ top information sources about BUCAS",
      "percent distribution", "inpatients", "none stated"),
    T("2.59b", "Percentage distribution of outpatient respondents’ top information sources about BUCAS",
      "percent distribution", "outpatients", "none stated"),
    T("2.60a", "Percentage distribution of inpatient respondents who can identify the benefits under BUCAS",
      "percent distribution", "inpatients", "none stated"),
    T("2.60b", "Percentage distribution of outpatient respondents who can identify the benefits under BUCAS",
      "percent distribution", "outpatients", "none stated"),
    T("2.61a", "Percentage distribution of inpatient respondents’ most commonly accessed services in BUCAS",
      "percent distribution", "inpatients", "none stated"),
    T("2.61b", "Percentage distribution of outpatient respondents’ most commonly accessed services in BUCAS",
      "percent distribution", "outpatients", "none stated"),
    T("2.62", "Percentage distribution of inpatient and outpatient respondents who have heard about GAMOT",
      "percent distribution", "inpatients+outpatients", "none stated"),
    T("2.63a", "Percentage distribution of inpatient respondents’ top information sources about GAMOT",
      "percent distribution", "inpatients", "none stated"),
    T("2.63b", "Percentage distribution of outpatient respondents’ top information sources about GAMOT",
      "percent distribution", "outpatients", "none stated"),
    T("2.64a", "Percentage distribution of inpatient respondents who can correctly identify the benefits under GAMOT",
      "percent distribution", "inpatients", "none stated"),
    T("2.64b", "Percentage distribution of outpatient respondents who can correctly identify the benefits under GAMOT",
      "percent distribution", "outpatients", "none stated"),
    T("2.65", "Descriptive statistics of socio-demographic characteristics of the inpatient and outpatient respondents",
      "descriptive stats", "inpatients+outpatients", "none stated"),

    # ------------------------------------------------------------------
    # Annex 3. Household-level indicators (F4) — Tables 3.1-3.36
    # ------------------------------------------------------------------
    T("3.1", "Percent of household respondents who have heard about UHC, by province",
      "percent", "households", "by province"),
    T("3.2", "Percent distribution of source of information about UHC of household respondents, by province",
      "percent distribution", "households", "by province"),
    T("3.3", "Percent distribution of household respondents who identify as registered to PhilHealth, by province",
      "percent distribution", "households", "by province"),
    T("3.4", "Percentage distribution of household respondents that are unregistered who know where to seek assistance for the registration process, by province",
      "percent distribution", "unregistered households", "by province"),
    T("3.5", "Percentage distribution of household respondents’ source of information on PhilHealth registration, by province",
      "percent distribution", "households", "by province"),
    T("3.6", "Percentage distribution of household respondents who can correctly list PhilHealth benefits under UHC reforms, by province",
      "percent distribution", "households", "by province"),
    T("3.7", "Percentage distribution of household respondents who have heard about YAKAP/Konsulta provider assignment under UHC, by province",
      "percent distribution", "households", "by province"),
    T("3.8", "Percentage distribution of household respondents who can identify the benefits under the PhilHealth YAKAP/Konsulta package",
      "percent distribution", "households", "none stated"),
    T("3.9", "Percentage distribution of household respondents’ top challenges in registration to PhilHealth for those who have undergone the registration process (including those who did not register successfully) in the past year, by province",
      "percent distribution", "households who have undergone the registration process", "by province"),
    T("3.10", "Percent distribution of household respondents by type of PhilHealth membership who self-report to have struggled to pay premiums in the past year, by province",
      "percent distribution", "household PhilHealth members", "by PhilHealth membership type × province"),
    T("3.11", "Percentage distribution of household respondents in HCPN that were referred from YAKAP/Konsulta provider or primary care provider, for most recent interaction, by province",
      "percent distribution", "households in HCPN", "by province"),
    T("3.12", "Percent distribution of household respondents that received assistance in making the appointment for that visit, by province",
      "percent distribution", "households", "by province"),
    T("3.13", "Percent distribution of household respondents that received assistance by discussing with them the different places they could visit to address their health problem, by province",
      "percent distribution", "households", "by province"),
    T("3.14", "Households' overall satisfaction rate with the referral system process for most recent interaction at the time of survey",
      "percent", "households", "none stated"),
    T("3.15", "Percentage distribution of household respondents who can correctly define a generic drug, by province",
      "percent distribution", "households", "by province"),
    T("3.16", "Percentage distribution of household respondents who chose a generic drug when buying medication, for last interaction, by province",
      "percent distribution", "households", "by province"),
    T("3.17", "Percentage distribution of household respondents’ reasons to purchase generic options (prescribed by doctor, cheaper, etc.) in the past 6 months, by province",
      "percent distribution", "households", "by province"),
    T("3.18", "Percentage distribution of household respondents’ reasons to purchase branded options (prescribed by doctor, cheaper, etc.) in the past 6 months, by province",
      "percent distribution", "households", "by province"),
    T("3.19", "Percent distribution of household respondents that accessed medications for free or subsidized, in the last 6 months, by province",
      "percent distribution", "households", "by province"),
    T("3.20", "Percent distribution of household respondents’ top sources of medicine subsidization (i.e., LGU, employer, etc.) in the past 6 months",
      "percent distribution", "households", "none stated"),
    T("3.21", "Percentage distribution of household respondents that have accessed alternate financial assistance program/s to help cover medical costs, for most recent interaction, out of all patients",
      "percent distribution", "households (all patients)", "none stated"),
    T("3.22", "Percentage distribution of household respondents that have accessed alternate financial assistance program/s to help cover medical costs, for most recent interaction, out of billed patients",
      "percent distribution", "households (billed patients)", "none stated"),
    T("3.23", "Percentage distribution of household respondents that experience incidence of reduced care due to financial reasons, for most recent interaction",
      "percent distribution", "households", "none stated"),
    T("3.24", "Percentage distribution of household respondents that experience incidence of forgone care due to financial reasons, for most recent interaction",
      "percent distribution", "households", "none stated",
      "printed with identical description to Table 3.25"),
    T("3.25", "Percentage distribution of household respondents that experience incidence of forgone care due to financial reasons, for most recent interaction",
      "percent distribution", "households", "none stated",
      "printed with identical description to Table 3.24"),
    T("3.26", "Percentage distribution of household respondents who have heard about ZBB",
      "percent distribution", "households", "none stated"),
    T("3.27", "Percentage distribution of household respondents’ top information sources about ZBB",
      "percent distribution", "households", "none stated"),
    T("3.28", "Percentage distribution of household respondents who can correctly identify the benefits under ZBB",
      "percent distribution", "households", "none stated"),
    T("3.29", "Percentage distribution of household respondents who have heard about BUCAS",
      "percent distribution", "households", "none stated"),
    T("3.30", "Percentage distribution of household respondents’ top information sources about BUCAS",
      "percent distribution", "households", "none stated"),
    T("3.31", "Percentage distribution of household respondents who can correctly identify the benefits under BUCAS",
      "percent distribution", "households", "none stated"),
    T("3.32", "Percentage distribution of household respondents’ mostly accessed services in BUCAS centers",
      "percent distribution", "households", "none stated"),
    T("3.33", "Percentage distribution of household respondents who have heard about GAMOT",
      "percent distribution", "households", "none stated"),
    T("3.34", "Percentage distribution of household respondents’ top information sources about GAMOT",
      "percent distribution", "households", "none stated"),
    T("3.35", "Percentage distribution of household respondents who can correctly identify the benefits under GAMOT",
      "percent distribution", "households", "none stated"),
    T("3.36", "Descriptive statistics of socio-demographic characteristics of the household sample",
      "descriptive stats", "households", "none stated"),

    # ------------------------------------------------------------------
    # Annex 4. Healthcare worker-level indicators (F2) — Tables 4.1-4.20
    # (The printed header carries a stray "Revised Table No." column artifact — ignored.)
    # ------------------------------------------------------------------
    T("4.1", "Number of healthcare workers per facility at point of survey, across level of facility and type of worker",
      "summary", "facilities", "by facility level × worker type"),
    T("4.2", "Average gap (%) in health workers per facility over the past year at the time of survey, across level of facility and type of worker",
      "average", "facilities", "by facility level × worker type"),
    T("4.3", "Percentage distribution of healthcare worker respondents who have ever heard about UHC",
      "percent distribution", HCW, "none stated"),
    T("4.4", "Percentage distribution of healthcare worker respondents who reported doing tasks that could be completed by a more junior staff",
      "percent distribution", HCW, "none stated",
      "printed with identical description to Table 4.5"),
    T("4.5", "Percentage distribution of healthcare worker respondents who reported doing tasks that could be completed by a more junior staff",
      "percent distribution", HCW, "none stated",
      "printed with identical description to Table 4.4"),
    T("4.6", "Percentage distribution of healthcare worker respondents' reasons for taking on tasks that should be performed by a different role",
      "percent distribution", HCW, "none stated"),
    T("4.7", "Percentage distribution of healthcare worker respondents’ views on Task Sharing across roles",
      "percent distribution", HCW, "none stated"),
    T("4.8", "Percentage distribution of healthcare worker respondents who reported receiving adequate support from their facilities to implement changes related to UHC, in the past year",
      "percent distribution", HCW, "none stated"),
    T("4.9", "Percentage distribution of healthcare worker respondents who reported receiving adequate support from their facilities to implement changes related to UHC, in the past year (continued)",
      "percent distribution", HCW, "none stated",
      "printed as (continued) of Table 4.8"),
    T("4.10", "Percentage distribution of healthcare worker respondents who would consider becoming YAKAP/Konsulta providers, if not already registered",
      "percent distribution", "HCWs not already registered as YAKAP/Konsulta providers", "none stated"),
    T("4.11", "Percentage distribution of healthcare worker respondents’ satisfaction on compensation, over the past 6 months",
      "percent distribution", HCW, "none stated"),
    T("4.12", "Percent of facilities who provide adequate professional development opportunities for healthcare workers, over the past 6 months",
      "percent", "facilities", "none stated"),
    T("4.13", "Percentage distribution of healthcare worker respondents’ satisfaction on available professional development/training opportunities, over the past 6 months",
      "percent distribution", HCW, "none stated"),
    T("4.14", "Percentage distribution of healthcare worker respondents rating on the intensity of workload based on amount of overtime, over the past 6 months",
      "percent distribution", HCW, "none stated"),
    T("4.15", "Percentage distribution of healthcare worker respondents’ satisfaction rating, over the past 6 months",
      "percent distribution", HCW, "none stated"),
    T("4.16", "Percentage distribution of healthcare worker respondents on top factors for job satisfaction and dissatisfaction, over the last 6 months",
      "percent distribution", HCW, "none stated"),
    T("4.17", "Percentage distribution of healthcare worker respondents who intended to leave current post in the next year",
      "percent distribution", HCW, "none stated"),
    T("4.18", "Percentage distribution of healthcare worker respondents on top reasons for healthcare workers who intend to leave in less than a year",
      "percent distribution", "HCWs who intend to leave in less than a year", "none stated"),
    T("4.19", "Percentage distribution of healthcare workers by role in the health facility",
      "percent distribution", HCW, "by role in the health facility"),
    T("4.20", "Percentage distribution of healthcare workers by specialization in the health facility",
      "percent distribution", HCW, "by specialization in the health facility"),
]

COLUMNS = [
    ("no", "Table No."),
    ("annex", "Annex"),
    ("instrument", "Instrument"),
    ("description", "Description (verbatim, SSRCS Form 1 §II-9)"),
    ("stat_type", "Stat Type"),
    ("universe", "Universe"),
    ("breakdown", "Breakdown"),
    ("weight", "Weight"),
    ("source_variables", "Source Variables"),
    ("mapping_status", "Mapping Status"),
    ("notes", "Notes"),
]

EXPECTED_COUNTS = {1: 36, 2: 105, 3: 36, 4: 20}


def validate():
    nos = [t["no"] for t in TABLES]
    assert len(nos) == len(set(nos)), "duplicate table numbers found"
    counts = {}
    for t in TABLES:
        counts[t["annex"]] = counts.get(t["annex"], 0) + 1
    for annex, expected in EXPECTED_COUNTS.items():
        got = counts.get(annex, 0)
        assert got == expected, f"Annex {annex}: expected {expected} rows, got {got}"
    return counts


def write_csv(path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([key for key, _ in COLUMNS])
        for t in TABLES:
            writer.writerow([t[key] for key, _ in COLUMNS])


def write_xlsx(path, counts):
    wb = Workbook()

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=DOH_GREEN)
    body_font = Font(name="Arial", size=10)
    title_font = Font(name="Arial", size=14, bold=True, color=DOH_GREEN)
    section_font = Font(name="Arial", size=11, bold=True, color=DOH_GREEN)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    # ---- README sheet (first, visible) ----
    rm = wb.active
    rm.title = "README"
    rm.sheet_view.showGridLines = False
    rm.column_dimensions["A"].width = 118

    def rm_row(row, text, font=body_font, wrap=True):
        c = rm.cell(row=row, column=1, value=text)
        c.font = font
        if wrap:
            c.alignment = wrap_top

    r = 1
    rm_row(r, "UHC Survey Year 2 — Tabulation Plan", title_font); r += 2
    rm_row(r, "SOURCE", section_font); r += 1
    rm_row(r, "PSA Statistical Survey Notification Form (SSRCS Form 1), §II item 9 "
              '"List of tables and other outputs to be generated", signed 2026-06-15. '
              "PDF: raw/Email-Ingest-2026-06-05/Attachments/2_DOH UHC Yr2_PSA SSRCS Form 1_signed.pdf "
              "(pages 3-9). Descriptions on the Plan sheet are transcribed verbatim, including "
              "printed quirks (2.49a/b duplicate 2.48a/b; 3.24/3.25 identical; 4.4/4.5 identical; "
              "1.21 and 4.9 '(continued)'; no plain Table 1.17 — only 1.17a/1.17b; Annex 4's printed "
              "header carries a stray 'Revised Table No.' column artifact, ignored)."); r += 2
    rm_row(r, "ROW COUNTS", section_font); r += 1
    for annex in sorted(counts):
        rm_row(r, f"{ANNEX_TITLE[annex]}: {counts[annex]} tables"); r += 1
    rm_row(r, f"TOTAL: {sum(counts.values())} tables"); r += 2
    rm_row(r, "WORKFLOW", section_font); r += 1
    rm_row(r, "1. This plan (generated by build_tabulation_plan.py — edit the script, never this "
              "workbook by hand)."); r += 1
    rm_row(r, "2. Variable mapping pass: fill Source Variables per table from the harmonization "
              "codebook; flip Mapping Status unmapped -> mapped."); r += 1
    rm_row(r, "3. Stata 12 do-files: one tabulation program per annex, emitting the committed "
              "tables via svy: tab / svy: proportion / svy: mean + tabout."); r += 2
    rm_row(r, "STATA 12 CONSTRAINT", section_font); r += 1
    rm_row(r, STATA12_CONSTRAINT); r += 1

    # ---- Plan sheet ----
    ws = wb.create_sheet("Plan")
    ws.freeze_panes = "A2"

    for col_idx, (_, header) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="center")

    for row_idx, t in enumerate(TABLES, start=2):
        for col_idx, (key, _) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=t[key])
            c.font = body_font
            c.border = border
            c.alignment = wrap_top

    # Column widths sized to content (caps keep the wide text columns wrapped).
    caps = {"description": 95, "universe": 42, "breakdown": 32, "notes": 42}
    for col_idx, (key, header) in enumerate(COLUMNS, start=1):
        longest = max([len(header)] + [len(str(t[key])) for t in TABLES])
        width = min(longest + 2, caps.get(key, 20))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(TABLES) + 1}"

    wb.save(path)


def apply_mappings():
    """Overlay variable mappings from mappings/annex*-mapping.csv onto TABLES.

    The SSRCS table list above is the immutable source of truth; the mapping
    CSVs are the evolving layer (one per annex, regenerable independently —
    this is what lets the per-annex mapping passes run in parallel without
    touching this script). Columns: no, source_variables, mapping_status,
    notes_add. mapping_status must be one of: mapped | partial | gap.
    """
    by_no = {t["no"]: t for t in TABLES}
    applied = 0
    for f in sorted((HERE / "mappings").glob("annex*-mapping.csv")):
        with f.open(encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                no = (row.get("no") or "").strip()
                t = by_no.get(no)
                if t is None:
                    print(f"  ! {f.name}: unknown table no {no!r} — skipped")
                    continue
                t["source_variables"] = (row.get("source_variables") or "").strip()
                status = (row.get("mapping_status") or "").strip()
                if status in {"mapped", "partial", "gap"}:
                    t["mapping_status"] = status
                extra = (row.get("notes_add") or "").strip()
                if extra:
                    t["notes"] = f"{t['notes']} | {extra}" if t["notes"] else extra
                applied += 1
        print(f"applied {f.name}")
    if applied:
        from collections import Counter
        status_counts = Counter(t["mapping_status"] for t in TABLES)
        print(f"mappings applied to {applied} rows — status: {dict(status_counts)}")


def main():
    counts = validate()
    apply_mappings()
    xlsx_path = HERE / "tabulation-plan.xlsx"
    csv_path = HERE / "tabulation-plan.csv"
    write_xlsx(xlsx_path, counts)
    write_csv(csv_path)
    for annex in sorted(counts):
        print(f"Annex {annex} ({ANNEX_INSTRUMENT[annex]}): {counts[annex]} tables")
    print(f"TOTAL: {sum(counts.values())} tables")
    print(f"Wrote: {xlsx_path}")
    print(f"Wrote: {csv_path}")


if __name__ == "__main__":
    main()
