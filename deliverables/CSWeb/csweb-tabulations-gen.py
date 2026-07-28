#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate the UHC Year 2 tabulation PREVIEW deliverables + catalog.

Hybrid preview engine (Carl, 2026-07-28; design spec
docs/superpowers/specs/2026-07-28-tabulations-preview-engine-design.md):

  Lane 1 (plan-driven) — for every cleanly-mapped plan row (~118), an unweighted
    pretest-only frequency of the mapped source variable, crossed by the
    committed breakdown, codes labeled as words. Each is "Preview of Table X.Y".
  Lane 2 (curated) — checkbox multi-code fields become multiple-response
    tallies; F2 answers are exploded from values_json and a curated handful of
    Annex-4 items tabulated.

Honest by construction: PRETEST phase only, UNWEIGHTED, every sheet stamped;
nothing here claims a committed weighted number (those come from the Stata lane).
Catastrophic-expenditure tables are marked `deferred`, not faked.

Outputs (all atomic; missing input -> exit nonzero, previous outputs stay):
  <out-dir>/UHC-Y2-Tabulations-<INST>-preview.xlsx  (gated data room)
  <data>/tabulations-preview.json   tidy cells keyed by table_no (gated)
  <data>/tabulations-manifest.json  per-table status + preview->table_no map (gated)
  <public>                          public catalog: counts + preview table_nos only

Cron: hourly ("7 * * * *"), flock /tmp/csweb-tabulations.lock, /opt/venvs/spss
(pandas + openpyxl). Off-box: pass --data-dir/--meta-dir/--out-dir/... to run
against pulled fixtures with no MySQL.
"""
import argparse
import csv
import datetime
import html
import json
import os
import re
import sys
import tempfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import tabulation_lib as T
import portal_shell as PS   # page chrome — /opt/portal_shell.py + portal.css on the box

REF_PHASE = "pretest"      # training/survey activity must never mix into previews
CAPI = ("f1", "f3", "f4")
INST_NAME = {"f1": "Facility Head (F1)", "f3": "Patient (F3)",
             "f4": "Household (F4)", "f2": "Health Care Worker (F2)"}

STAMP = ("PRETEST PREVIEW — unweighted frequencies from live sync data, phase="
         + REF_PHASE + " cases only. NOT the official PSA-committed tables "
         "(those come from the weighted Stata lane).")
DEFERRED_NOTE = ("derived indicator — deferred to the weighted lane / CTP "
                 "catastrophic-expenditure alignment; not computed here")


def defaults():
    p = argparse.ArgumentParser(description="UHC Y2 tabulation preview generator")
    p.add_argument("--plan", default="/opt/tabulation-plan.csv")
    p.add_argument("--data-dir", default="/opt/app/lamp/www/docs/data")
    p.add_argument("--meta-dir", default="/opt/spss-meta")
    p.add_argument("--out-dir", default=None, help="default <data-dir>/tabulations")
    p.add_argument("--manifest", default=None,
                   help="default <data-dir>/tabulations-manifest.json")
    p.add_argument("--preview-json", default=None,
                   help="default <data-dir>/tabulations-preview.json")
    p.add_argument("--public",
                   default="/opt/app/capi-www/projects/uhc-y2/tabulations.json")
    p.add_argument("--page",
                   default="/opt/app/capi-www/projects/uhc-y2/tabulations/index.html",
                   help="the live portal Tabulations page (this generator owns it; "
                        "build_portal.py no longer writes it)")
    a = p.parse_args()
    a.out_dir = a.out_dir or os.path.join(a.data_dir, "tabulations")
    a.manifest = a.manifest or os.path.join(a.data_dir, "tabulations-manifest.json")
    a.preview_json = a.preview_json or os.path.join(a.data_dir, "tabulations-preview.json")
    return a


def load_pretest(csv_path):
    """Responses CSV as string DataFrame, filtered to the reference phase."""
    df = pd.read_csv(csv_path, dtype=str, keep_default_na=False, na_values=[])
    if "phase" in df.columns:
        df = df[df["phase"] == REF_PHASE]
    return df


def is_catastrophic(desc):
    return "catastrophic" in (desc or "").lower()


def _san_sheet(no, used):
    """Excel-legal, unique sheet name from a table number."""
    name = "T" + re.sub(r"[\[\]:*?/\\]", "", str(no))[:29]
    base, i = name, 2
    while name.lower() in used:
        suf = "-%d" % i
        name = base[:31 - len(suf)] + suf
        i += 1
    used.add(name.lower())
    return name


def build_tables(a):
    """Return (tables, plan_len). Each table dict:
    {no, inst, annex, title, breakdown, note, universe, status, kind, cells, n}."""
    plan = list(csv.DictReader(open(a.plan, encoding="utf-8-sig")))
    headers = {i: T.read_headers(os.path.join(a.data_dir, "%s_responses.csv" % i))
               for i in CAPI}
    rows = T.classify_rows(plan, headers)
    vlabels = T.load_value_labels(a.meta_dir)          # {inst:{col:{code:label}}}

    data = {i: load_pretest(os.path.join(a.data_dir, "%s_responses.csv" % i)) for i in CAPI}

    tables = []
    for r in rows:
        no, inst, klass = r["no"], r["inst"], r["klass"]
        base = {"no": no, "inst": inst, "annex": r["annex"],
                "title": r["description"], "breakdown": r["breakdown"],
                "universe": r["universe"], "note": "", "kind": "freq",
                "status": "partial", "cells": [], "n": 0}

        if is_catastrophic(r["description"]):          # honesty: never fake this
            base.update(status="deferred", note=DEFERRED_NOTE)
            tables.append(base); continue

        if klass == "gap":
            base["status"] = "gap"; tables.append(base); continue
        if klass == "partial":
            base["status"] = "partial"; tables.append(base); continue
        if klass == "f2":
            tables.append(base); continue              # F2 handled in build_f2

        df = data[inst]
        vlab = vlabels.get(inst, {})
        if klass == "previewable":
            bdcol, bnote = T.resolve_breakdown(inst, r["breakdown"], headers[inst])
            cells = T.tabulate(df, r["col"], vlab.get(r["col"], {}),
                               breakdown_col=bdcol,
                               breakdown_vlabels=vlab.get(bdcol, {}) if bdcol else None)
            base.update(kind="freq", cells=cells, n=len(df), status="preview",
                        note=bnote or "")
            tables.append(base)
        elif klass == "multi":
            if not r["col"] or r["col"] not in df.columns:
                base["status"] = "partial"; tables.append(base); continue
            cells = T.multi_tally(df, r["col"], vlab.get(r["col"], {}))
            n_resp = next((c["n"] for c in cells if c.get("meta") == "respondents"), 0)
            base.update(kind="multi", cells=cells, n=n_resp, status="preview",
                        note="multiple-response — percentages are of respondents and sum > 100")
            tables.append(base)
    return tables, len(plan)


def build_f2(a, tables):
    """Fill the curated F2 previews (Annex 4) in place; return the F2 table list."""
    f2csv = os.path.join(a.data_dir, "f2_responses.csv")
    f2lab = json.load(open(os.path.join(a.meta_dir, "f2-item-labels.json"), encoding="utf-8"))
    rows = list(csv.DictReader(open(f2csv, encoding="utf-8-sig")))
    df, _ = T.explode_f2(rows, f2lab)
    by_no = {t["no"]: t for t in tables if t["inst"] == "f2"}
    for item, no, fallback in T.F2_PREVIEWS:
        t = by_no.get(no)
        if t is None:                                  # plan row absent — synthesize
            t = {"no": no, "inst": "f2", "annex": "4", "title": fallback,
                 "breakdown": "", "universe": "", "kind": "f2", "cells": [], "n": 0}
            tables.append(t)
        if item in df.columns:
            cells = T.tabulate(df, item)               # F2 answers are English text
            t.update(kind="f2", cells=cells, n=len(df), status="preview",
                     note="all mirrored F2 submissions (no phase split)")
        else:
            t.update(status="partial", note="F2 item %s not present in current data" % item)
    return [t for t in tables if t["inst"] == "f2"]


# ------------------------------------------------------------- packaging -----

def _hdr(kind):
    if kind == "multi":
        return ["Option", "n", "% of respondents"]
    return ["Group", "Category", "n", "%"]


def write_workbook(inst, tables, path, gen_ts):
    """One workbook: TOC + a sheet per previewed table."""
    previews = [t for t in tables if t["status"] == "preview" and t["cells"]]
    wb = Workbook()
    toc = wb.active
    toc.title = "TOC"
    toc["A1"] = "UHC Survey Year 2 — Tabulations (preview): " + INST_NAME[inst]
    toc["A1"].font = Font(bold=True, size=13)
    toc["A2"] = "Generated %s · %d preview tables" % (gen_ts, len(previews))
    toc["A3"] = STAMP
    toc["A3"].font = Font(italic=True, size=9)
    hr = 5
    for j, h in enumerate(["Table", "Title", "Kind", "n", "Sheet"], start=1):
        c = toc.cell(hr, j, h); c.font = Font(bold=True)
    used = set()
    for i, t in enumerate(previews):
        sn = _san_sheet(t["no"], used)
        row = hr + 1 + i
        toc.cell(row, 1, t["no"])
        toc.cell(row, 2, t["title"])
        toc.cell(row, 3, {"freq": "frequency", "multi": "multi-response", "f2": "F2"}[t["kind"]])
        toc.cell(row, 4, t["n"])
        link = toc.cell(row, 5, sn)
        link.hyperlink = "#'%s'!A1" % sn
        link.font = Font(color="0563C1", underline="single")
        # --- the table sheet ---
        ws = wb.create_sheet(sn)
        ws["A1"] = "Preview of Table %s — %s" % (t["no"], t["title"])
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = STAMP
        ws["A2"].font = Font(italic=True, size=9)
        line = 3
        for extra in (("Universe: " + t["universe"]) if t["universe"] else "", t["note"]):
            if extra:
                ws.cell(line, 1, extra).font = Font(italic=True, size=9)
                line += 1
        line += 1
        hdr = _hdr(t["kind"])
        for j, h in enumerate(hdr, start=1):
            ws.cell(line, j, h).font = Font(bold=True)
        line += 1
        for cell in t["cells"]:
            if cell.get("meta") == "respondents":
                continue
            if t["kind"] == "multi":
                ws.cell(line, 1, cell["category"]); ws.cell(line, 2, cell["n"])
                ws.cell(line, 3, cell["pct_resp"])
            else:
                ws.cell(line, 1, cell.get("group", "")); ws.cell(line, 2, cell["category"])
                ws.cell(line, 3, cell["n"]); ws.cell(line, 4, cell["pct"])
            line += 1
        if t["kind"] == "multi":
            nr = next((c["n"] for c in t["cells"] if c.get("meta") == "respondents"), 0)
            ws.cell(line + 1, 1, "Respondents (base): %d" % nr).font = Font(italic=True, size=9)
        for col in range(1, len(hdr) + 1):              # rough autofit
            ws.column_dimensions[get_column_letter(col)].width = 46 if col <= 2 else 14
    if not previews:                                    # never leave an empty book
        toc.cell(hr + 1, 1, "No previews available for this instrument yet.")
    _atomic_save(wb, path)


def _atomic_save(wb, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=os.path.dirname(path), suffix=".xlsx")
    os.close(fd)
    wb.save(tmp)
    os.replace(tmp, path)
    os.chmod(path, 0o644)


def _atomic_json(path, obj, compact=False):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=os.path.dirname(path), suffix=".tmp")
    with os.fdopen(fd, "w", encoding="utf-8") as fh:
        json.dump(obj, fh, separators=(",", ":") if compact else None,
                  indent=None if compact else 1, ensure_ascii=False)
    os.replace(tmp, path)
    os.chmod(path, 0o644)


# ------------------------------------------------------------- portal page ---

ANNEXES = [("1", "f1", "Annex 1 — Facility level (F1)"),
           ("2", "f3", "Annex 2 — Patient level (F3)"),
           ("3", "f4", "Annex 3 — Household level (F4)"),
           ("4", "f2", "Annex 4 — Health care worker (F2)")]

STATUS_CHIP = {                     # status -> (chip css class, label)
    "preview":  ("st-prev", "preview ready"),
    "partial":  ("st-plan", "planned"),
    "gap":      ("st-plan", "planned · no source item"),
    "deferred": ("st-defer", "deferred · weighted lane"),
}

_PAGE_CSS = """
.tb-kpis{display:flex;gap:14px;flex-wrap:wrap;margin:14px 0 18px}
.tb-kpi{background:var(--surface,#fff);border:1px solid var(--line,#d9e2dc);border-radius:12px;padding:12px 18px;min-width:130px}
.tb-kpi b{display:block;font-size:24px;color:var(--verde-900,#04331d)}
.tb-kpi span{font-size:11.5px;color:var(--ink-3,#5c6b63);text-transform:uppercase;letter-spacing:.06em}
.tb-stamp{background:#fdf7e8;border:1px solid #f0dcae;color:#6b5418;border-radius:10px;padding:10px 14px;font-size:12.5px;margin:0 0 16px;line-height:1.5}
.tb-tools{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin:0 0 14px}
.tb-tools input[type=search]{flex:1;min-width:220px;font:13.5px inherit;padding:8px 12px;border:1px solid var(--line-strong,#c3d0c8);border-radius:9px}
.tb-chip{font:600 12px inherit;padding:6px 12px;border-radius:20px;border:1px solid var(--line-strong,#c3d0c8);background:#fff;cursor:pointer;color:var(--ink-2,#33413a)}
.tb-chip.on{background:var(--verde,#046a38);border-color:var(--verde,#046a38);color:#fff}
.tb-dl{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:12px;margin:0 0 20px}
.tb-dlcard{background:var(--surface,#fff);border:1px solid var(--line,#d9e2dc);border-radius:12px;padding:13px 16px;font-size:13px}
.tb-dlcard b{color:var(--verde-900,#04331d)}
.tb-dlcard .m{color:var(--ink-3,#5c6b63);font-size:11.5px;margin:3px 0 8px}
.tb-dlcard a{font-weight:700;color:var(--verde,#046a38);text-decoration:none}
.tb-annex{border:1px solid var(--line,#d9e2dc);border-radius:12px;margin-bottom:12px;background:var(--surface,#fff);overflow:hidden}
.tb-annex summary{cursor:pointer;padding:13px 18px;font-weight:700;color:var(--verde-900,#04331d);display:flex;justify-content:space-between;align-items:center;gap:10px;list-style:none}
.tb-annex summary::-webkit-details-marker{display:none}
.tb-annex summary .cnt{font-weight:600;font-size:12px;color:var(--ink-3,#5c6b63)}
.tb-annex[open] summary{border-bottom:1px solid var(--line,#d9e2dc)}
.trow{display:flex;gap:12px;align-items:baseline;padding:9px 18px;border-top:1px solid #eef3f0;font-size:13.5px}
.trow:first-of-type{border-top:0}
.trow .no{font-family:ui-monospace,Consolas,monospace;font-weight:700;color:var(--verde,#046a38);min-width:52px}
.trow .tt{flex:1;line-height:1.45}
.trow .st{font-size:10.5px;font-weight:700;letter-spacing:.05em;text-transform:uppercase;padding:3px 9px;border-radius:6px;white-space:nowrap}
.st-prev{background:#eefaf3;color:#047857;border:1px solid #b9e2cb}
.st-plan{background:#f1f5f9;color:#64748b;border:1px solid #dde5ec}
.st-defer{background:#fdf7e8;color:#a16207;border:1px solid #f0dcae}
.tb-view{font:700 11.5px inherit;color:var(--verde,#046a38);background:none;border:1px solid var(--verde,#046a38);border-radius:7px;padding:4px 10px;cursor:pointer;white-space:nowrap}
.tb-view:hover{background:var(--verde,#046a38);color:#fff}
.tprev{display:none;padding:6px 18px 16px 82px}
.tprev.open{display:block}
.tprev table{border-collapse:collapse;font-size:12.5px;min-width:340px}
.tprev th,.tprev td{border:1px solid var(--line,#d9e2dc);padding:5px 10px;text-align:left}
.tprev th{background:#f2f6f3;color:var(--verde-900,#04331d)}
.tprev .pn{font-size:11px;color:var(--ink-3,#5c6b63);margin:6px 0 0}
.tprev .perr{font-size:12.5px;color:#a16207;background:#fdf7e8;border:1px solid #f0dcae;border-radius:8px;padding:8px 12px}
.trow.hide,.tb-annex.hide{display:none}
@media (max-width:700px){.tprev{padding-left:18px}.trow{flex-wrap:wrap}}
"""

_PAGE_JS = """
(function(){
  var q=document.getElementById('tbq'), chips=[].slice.call(document.querySelectorAll('.tb-chip'));
  var rows=[].slice.call(document.querySelectorAll('.trow'));
  var annexes=[].slice.call(document.querySelectorAll('.tb-annex'));
  var stFilter='all';
  function apply(){
    var t=(q.value||'').toLowerCase();
    annexes.forEach(function(ax){
      var vis=0;
      [].slice.call(ax.querySelectorAll('.trow')).forEach(function(r){
        var okT=!t || r.getAttribute('data-k').indexOf(t)>=0;
        var okS=stFilter==='all' || r.getAttribute('data-st')===stFilter;
        var show=okT&&okS;
        r.classList.toggle('hide',!show); if(show)vis++;
      });
      ax.classList.toggle('hide',vis===0);
      if((t||stFilter!=='all')&&vis>0) ax.setAttribute('open','');
    });
  }
  q.addEventListener('input',apply);
  chips.forEach(function(c){c.addEventListener('click',function(){
    stFilter=c.getAttribute('data-f');
    chips.forEach(function(x){x.classList.toggle('on',x===c)});
    apply();
  });});

  // lazy preview fetch: the numbers live in the GATED data room (staff tier).
  var prevP=null;
  function getPreviews(){
    if(!prevP){ prevP=fetch('/docs/data/tabulations-preview.json',{credentials:'same-origin'})
      .then(function(r){ if(!r.ok) throw r.status; return r.json(); }); }
    return prevP;
  }
  function td(tr,txt,th){var c=document.createElement(th?'th':'td');c.textContent=txt;tr.appendChild(c);}
  function render(panel,rows,kind,note){
    panel.textContent='';
    var tb=document.createElement('table'),hr=document.createElement('tr');
    var multi=(kind==='multi');
    (multi?['Option','n','% of respondents']:['Group','Category','n','%']).forEach(function(h){td(hr,h,1)});
    tb.appendChild(hr);
    rows.forEach(function(r){
      var tr=document.createElement('tr');
      if(multi){ td(tr,r.category); td(tr,String(r.n)); td(tr,String(r.pct)); }
      else{ td(tr,r.group||'—'); td(tr,r.category); td(tr,String(r.n)); td(tr,String(r.pct)); }
      tb.appendChild(tr);
    });
    panel.appendChild(tb);
    var p=document.createElement('div');p.className='pn';
    p.textContent=(note?note+' · ':'')+'Unweighted pretest preview — not an official weighted table.';
    panel.appendChild(p);
  }
  [].slice.call(document.querySelectorAll('.tb-view')).forEach(function(btn){
    btn.addEventListener('click',function(){
      var no=btn.getAttribute('data-no');
      var panel=document.getElementById('tp-'+no.replace(/[^A-Za-z0-9.]/g,''));
      if(!panel) return;
      if(panel.classList.contains('open')){ panel.classList.remove('open'); btn.textContent='View preview'; return; }
      panel.classList.add('open'); btn.textContent='Hide';
      if(panel.getAttribute('data-done')) return;
      panel.textContent='Loading…';
      getPreviews().then(function(j){
        var mine=(j.rows||[]).filter(function(r){return r.table_no===no;});
        if(!mine.length){ panel.innerHTML='<div class="perr">No preview cells found for this table.</div>'; }
        else{ render(panel,mine,mine[0].kind,mine[0].note||''); }
        panel.setAttribute('data-done','1');
      }).catch(function(e){
        panel.innerHTML='<div class="perr">'+(e===401
          ?'Preview numbers require a <b>staff</b> sign-in (data-room access). The catalog and Excel downloads above follow the same access tiers.'
          :'Could not load preview numbers ('+e+'). Try again shortly.')+'</div>';
        prevP=null;
      });
    });
  });
})();
"""


def _pid(no):
    """Panel id from a table number (matches the JS sanitizer)."""
    return re.sub(r"[^A-Za-z0-9.]", "", str(no))


def build_page(tables, planned, files, status_counts, gen_ts, path):
    """The live portal Tabulations page — owned by this generator (hourly cron),
    so the KPI numbers can never go stale the way the static build's did."""
    e = html.escape
    n_prev = status_counts.get("preview", 0)

    kpis = ('<div class="tb-kpis">'
            '<div class="tb-kpi"><b>%d</b><span>tables planned</span></div>'
            '<div class="tb-kpi"><b>%d</b><span>preview ready</span></div>'
            '<div class="tb-kpi"><b>0</b><span>built (official)</span></div>'
            '<div class="tb-kpi"><b>hourly</b><span>refresh</span></div></div>'
            % (planned, n_prev))

    stamp = ('<div class="tb-stamp"><b>Previews are PRETEST / UNWEIGHTED.</b> %s '
             'The official tables are produced in the weighted Stata lane.</div>'
             % e(STAMP))

    tools = ('<div class="tb-tools"><input id="tbq" type="search" '
             'placeholder="Search table number or title…" aria-label="Search tables">'
             '<button class="tb-chip on" data-f="all">All</button>'
             '<button class="tb-chip" data-f="preview">Preview ready</button>'
             '<button class="tb-chip" data-f="partial">Planned</button>'
             '<button class="tb-chip" data-f="deferred">Deferred</button></div>')

    dl = ['<div class="tb-dl">']
    for f in files:
        if f.get("previews"):
            dl.append('<div class="tb-dlcard"><b>%s</b>'
                      '<div class="m">%d preview tables · .xlsx · sign-in required</div>'
                      '<a href="/docs/data/%s">Download workbook</a></div>'
                      % (e(INST_NAME.get(f["instrument"].lower(), f["instrument"])),
                         f["previews"], e(f["file"])))
    dl.append('<div class="tb-dlcard"><b>Everything else</b>'
              '<div class="m">case exports, codebook, CSPro apps</div>'
              '<a href="/docs/data/">Open the data room</a></div></div>')

    by_no = {}
    for t in tables:
        by_no.setdefault(t["inst"], []).append(t)
    ax_html = []
    for annex, inst, title in ANNEXES:
        rows = by_no.get(inst, [])
        prev_ct = sum(1 for t in rows if t["status"] == "preview")
        body = []
        for t in rows:
            cls, label = STATUS_CHIP.get(t["status"], ("st-plan", t["status"]))
            key = e((t["no"] + " " + t["title"]).lower(), quote=True)
            view = ('<button class="tb-view" data-no="%s">View preview</button>'
                    % e(t["no"], quote=True)) if t["status"] == "preview" else ""
            body.append('<div class="trow" data-k="%s" data-st="%s">'
                        '<span class="no">%s</span><span class="tt">%s</span>'
                        '<span class="st %s">%s</span>%s</div>'
                        % (key, e(t["status"], quote=True), e(t["no"]), e(t["title"]),
                           cls, label, view))
            if t["status"] == "preview":
                body.append('<div class="tprev" id="tp-%s"></div>' % _pid(t["no"]))
        ax_html.append('<details class="tb-annex"%s><summary>%s'
                       '<span class="cnt">%d tables · %d preview</span></summary>%s</details>'
                       % (" open" if annex == "1" else "", e(title), len(rows),
                          prev_ct, "".join(body)))

    head = ('<div class="page-head"><div class="eyebrow">UHC Survey Year 2</div>'
            '<h1>Tabulations</h1><p class="lead">The full PSA-committed output catalog '
            '(%d tables), with unweighted pretest previews for every table the live data '
            'can already fill. Preview numbers and Excel downloads follow the console '
            'access tiers.</p></div>' % planned)

    page = (PS.open_shell("Tabulations — UHC Survey Y2",
                          "PSA-committed tabulation catalog with live pretest previews.",
                          active=PS.P + "/tabulations/",
                          crumbs=[("UHC Survey Year 2", PS.P + "/"), ("Tabulations", None)],
                          tb_right=PS.PILL_LIVE, extra_css=_PAGE_CSS)
            + head + kpis + stamp + tools + "".join(dl) + "".join(ax_html)
            + PS.close_shell(
                footer_html="Generated %s · refreshes hourly from synced instrument data."
                            % e(gen_ts),
                body_extra="<script>%s</script>" % _PAGE_JS))

    os.makedirs(os.path.dirname(path), exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=os.path.dirname(path), suffix=".tmp")
    with os.fdopen(fd, "w", encoding="utf-8") as fh:
        fh.write(page)
    os.replace(tmp, path)
    os.chmod(path, 0o644)


def tidy_rows(tables):
    """Flatten previewed cells into one row per table x category x group."""
    out = []
    for t in tables:
        if t["status"] != "preview":
            continue
        for cell in t["cells"]:
            if cell.get("meta") == "respondents":
                continue
            out.append({"table_no": t["no"], "instrument": t["inst"].upper(),
                        "annex": t["annex"], "kind": t["kind"], "title": t["title"],
                        "breakdown": t["breakdown"], "note": t["note"],
                        "universe": t["universe"],
                        "group": cell.get("group", ""), "category": cell["category"],
                        "n": cell["n"],
                        "pct": cell.get("pct", cell.get("pct_resp", 0.0)),
                        "status": "preview"})
    return out


def main():
    a = defaults()
    required = [a.plan] + [os.path.join(a.data_dir, "%s_responses.csv" % i)
                           for i in CAPI] + \
        [os.path.join(a.data_dir, "f2_responses.csv"),
         os.path.join(a.meta_dir, "f2-item-labels.json")]
    for f in required:
        if not os.path.exists(f):
            sys.exit("missing input: %s (keeping previous outputs)" % f)

    now = datetime.datetime.now(datetime.timezone.utc)
    gen_ts = now.strftime("%Y-%m-%d %H:%M UTC")

    tables, planned = build_tables(a)
    build_f2(a, tables)

    # workbooks per instrument (F1/F3/F4/F2)
    files = []
    for inst in ("f1", "f3", "f4", "f2"):
        itabs = [t for t in tables if t["inst"] == inst]
        name = "UHC-Y2-Tabulations-%s-preview.xlsx" % inst.upper()
        path = os.path.join(a.out_dir, name)
        write_workbook(inst, itabs, path, gen_ts)
        prev = [t for t in itabs if t["status"] == "preview" and t["cells"]]
        files.append({"file": "tabulations/" + name, "instrument": inst.upper(),
                      "previews": len(prev), "bytes": os.path.getsize(path)})

    # tidy preview JSON (feeds the Phase-2 in-page viewer)
    _atomic_json(a.preview_json,
                 {"generated": gen_ts, "stamp": STAMP, "rows": tidy_rows(tables)})

    # status roll-up + per-table records
    status_counts, by_annex = {}, {}
    table_recs, preview_nos = [], []
    for t in tables:
        status_counts[t["status"]] = status_counts.get(t["status"], 0) + 1
        by_annex.setdefault(t["inst"].upper(), 0)
        by_annex[t["inst"].upper()] += 1
        table_recs.append({"no": t["no"], "instrument": t["inst"].upper(),
                           "annex": t["annex"], "kind": t["kind"],
                           "title": t["title"], "status": t["status"]})
        if t["status"] == "preview":
            preview_nos.append(t["no"])
    n_preview = status_counts.get("preview", 0)

    _atomic_json(a.manifest, {
        "generated": gen_ts, "stamp": STAMP, "planned": planned,
        "preview": n_preview, "status_counts": status_counts,
        "files": files, "tables": table_recs,
        "preview_table_nos": sorted(preview_nos)})

    # public catalog — counts + which table_nos have a preview; NO cell numbers
    preview_by_annex = {}
    for t in tables:
        if t["status"] == "preview":
            preview_by_annex[t["inst"].upper()] = preview_by_annex.get(t["inst"].upper(), 0) + 1
    _atomic_json(a.public, {
        "generated": gen_ts, "planned": planned, "built": 0, "preview": n_preview,
        "by_annex": by_annex, "preview_by_annex": preview_by_annex,
        "status_counts": status_counts, "preview_table_nos": sorted(preview_nos)},
        compact=True)

    # the live portal page — this generator owns it (fresh KPIs every run)
    build_page(tables, planned, files, status_counts, gen_ts, a.page)

    print("wrote %d workbooks + portal page | previews=%d planned=%d | status=%s"
          % (len(files), n_preview, planned, status_counts))


if __name__ == "__main__":
    main()
