#!/usr/bin/env python3
r"""Generate the OFFICIAL per-instrument codebooks for UHC Survey Year 2.

Decision (Carl, 2026-07-21, options diagram codebook-standard-options-2026-07-21):
Option A now — a codebook GENERATED from the build artifacts every time, so it
can never drift from the deployed instruments (retires the manual CSPro
Dictionary-Macros xlsx export) — with Option B (DDI 2.5 packaging via the
World Bank Metadata Editor -> PSADA/NADA) at dissemination time. Content is
DDI-mapped (every column corresponds to a DDI-Codebook 2.5 variable element),
presentation is DHS-recode-manual-style tables, and the documentation format
matches how PSA documents its own surveys (NDHS 2022 = NADA/DDI catalog).
DHS variable RENAMING is deliberately NOT adopted (names stay questionnaire-
traceable); DHS missing-code conventions are adopted where cheap (#743).

Inputs (repo checkout, committed files only — safe in worktrees):
  CSPro/{F1,F3,F4}/<Base>.dcf        names, labels, types, value sets (truth)
  CSPro/{F1,F3,F4}/<Base>.ent.qsf    literal question text (EN; YAML, 7 langs)
  CSPro/versions.json                per-instrument build versions (SSOT)
  CSWeb/f2-item-labels.json          F2 (PWA) item labels + types

Outputs (--out-dir, default codebook-out/ next to this script; REPO-ONLY for
now — Carl has not yet decided on data-room publishing):
  F1-Codebook-v<ver>.xlsx            cover sheet + DHS-style variables table
  F3-Codebook-v<ver>.xlsx            (same per instrument)
  F4-Codebook-v<ver>.xlsx
  F2-Codebook-<date>.xlsx            PWA instrument (no CSPro version line)
  UHC-Y2-Codebook-<date>.html        all four instruments, printable (print
                                     to PDF from the browser for a PDF cut)

DDI column mapping (for the later Metadata Editor import stage):
  Variable -> var/@name · Label -> var/labl · Question (EN) -> var/qstn
  Values -> var/catgry · Special codes -> missing conventions · Type/Width ->
  var/@dcml + varFormat · Record/Roster -> fileStr/recGrp context.

Run:  py deliverables/data-harmonization/generate_codebook.py
"""
import argparse, datetime, html, json, os, re, sys

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DELIV = os.path.abspath(os.path.join(HERE, ".."))
CSPRO = os.path.join(DELIV, "CSPro")
BASE = {"f1": ("F1", "FacilityHeadSurvey"),
        "f3": ("F3", "PatientSurvey"),
        "f4": ("F4", "HouseholdSurvey")}
NAMES = {"f1": "F1 — Facility Head Survey", "f3": "F3 — Patient Survey",
         "f4": "F4 — Household Survey", "f2": "F2 — Healthcare Worker Survey (PWA)"}
SPECIAL_RE = re.compile(r"don.?t\s*know|refused|no answer|not applicable|declined", re.I)

GREEN, DARKGREEN, GOLD, INKLIGHT = "006B3F", "004D2C", "E5B23B", "EEF3F0"
THIN = Side(style="thin", color="B7C4BD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _en(o):
    for l in (o.get("labels") or []):
        if l.get("language") == "EN" and l.get("text"):
            return l["text"]
    for l in (o.get("labels") or []):
        if l.get("text"):
            return l["text"]
    return ""


def strip_html(s):
    s = re.sub(r"<[^>]+>", " ", s or "")
    s = html.unescape(s)
    return re.sub(r"\s+", " ", s).strip()


def load_qsf(path):
    """UPPER item name -> literal EN question text ('' if unlisted)."""
    d = yaml.safe_load(open(path, encoding="utf-8-sig"))
    out = {}
    for q in d.get("questions", []):
        name = (q.get("name") or "").split(".")[-1].upper()
        texts = []
        for c in q.get("conditions", []):
            t = strip_html((c.get("questionText") or {}).get("EN", ""))
            if t and t not in texts:
                texts.append(t)
        if name and texts:
            out[name] = texts[0] + (" [+%d conditional variant(s)]" % (len(texts) - 1)
                                    if len(texts) > 1 else "")
    return out


def parse_dcf(path):
    """Ordered item rows straight from the dictionary (ids first, dcf order)."""
    d = json.load(open(path, encoding="utf-8"))
    rows = []
    for lv in d.get("levels", []):
        blocks = []
        if lv.get("ids"):
            blocks.append(("(case ids)", "Case identifiers", 1, lv["ids"].get("items", [])))
        for r in lv.get("records", []):
            occ = ((r.get("occurrences") or {}).get("maximum") or 1)
            blocks.append((r.get("name", ""), _en(r) or r.get("name", ""), occ,
                           r.get("items", [])))
        for rec_name, rec_label, rec_occ, items in blocks:
            for it in items:
                vals, specials = [], []
                vs = it.get("valueSets") or []
                if vs:
                    for val in vs[0].get("values", []):
                        lab = _en(val)
                        for p in val.get("pairs", []):
                            if "from" in p:
                                vals.append("%s–%s%s" % (p["from"], p.get("to", ""),
                                                         (" = " + lab) if lab else ""))
                            elif "value" in p:
                                code = str(p["value"]).strip()
                                entry = "%s = %s" % (code, lab or code)
                                (specials if SPECIAL_RE.search(lab or "")
                                 or code in ("-98", "-99") else vals).append(entry)
                item_occ = ((it.get("occurrences") or {}).get("maximum") or 1)
                rows.append({
                    "name": it.get("name", "").lower(),
                    "label": _en(it),
                    "record": rec_label + (" (roster ×%d)" % rec_occ if rec_occ > 1 else ""),
                    "type": ("numeric" if it.get("contentType") == "numeric" else
                             it.get("contentType", "alpha")),
                    "width": "%s%s" % (it.get("length", ""),
                                       ("." + str(it["decimalPlaces"]))
                                       if it.get("decimalPlaces") else ""),
                    "occ": item_occ,
                    "values": vals, "specials": specials,
                    "uname": it.get("name", "").upper()})
    return rows


SKIP_ROW = re.compile(r"^\|\s*Q(\d+)\s+([A-Z][A-Z0-9_]*)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*$",
                      re.M)
VAL_ROW = re.compile(r"^\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*$", re.M)
SEVERITY = re.compile(r"HARD|SOFT|GATE|^[—–-]+")


def parse_logic_doc(path):
    """Skip rules + validation rows from F<i>-Skip-Logic-and-Validations.md.

    Shape-based, not section-scoped (the three docs number sections
    differently): a skip rule is any 3-col row '| Qn SUFFIX | cond | skip-to |';
    a validation is any 3-col row whose last cell is a severity (HARD/SOFT/
    GATE/—). §1's 4-column disposition tables match neither."""
    text = open(path, encoding="utf-8").read().replace("**", "")
    skips = [(int(m.group(1)), m.group(2), m.group(3).strip(), m.group(4).strip())
             for m in SKIP_ROW.finditer(text)]
    vrows = []
    for m in VAL_ROW.finditer(text):
        item, rule, sev = (m.group(i).strip() for i in (1, 2, 3))
        if item in ("Item", "Q", "---") or set(item) <= {"-"}:
            continue
        if SEVERITY.search(sev):
            vrows.append((item, rule, sev))
    return skips, vrows


def logic_maps(skips, vrows, rows):
    """Per-variable universe / routing / validation strings + coverage stats."""
    byname = {it["uname"]: it for it in rows}
    qidx = {}
    for it in rows:
        m = re.match(r"Q(\d+)_", it["uname"])
        if m:
            qidx.setdefault(int(m.group(1)), []).append(it["uname"])
    routing, universe, unmatched = {}, {}, []
    matched = 0
    for qn, suf, cond, skipto in skips:
        if "no skip" in skipto.lower():
            continue
        uname = "Q%d_%s" % (qn, suf)
        if uname in byname:
            matched += 1
            routing.setdefault(uname, []).append("%s → %s" % (cond, skipto))
        else:
            # a source absent from the dcf = a superseded/stale spec row (e.g. F4's
            # flat Section N) — never derive universe from it either
            unmatched.append(uname)
            continue
        m = re.search(r"skip\s+([^);]*)", skipto, re.I)     # "… (skip Q14, Q15; …)"
        if m:
            for qt in re.findall(r"Q(\d+)", m.group(1)):
                for u in qidx.get(int(qt), []):
                    universe.setdefault(u, []).append(
                        "not asked when Q%d %s %s" % (qn, suf, cond))
    valmap = {}
    for item, rule, sev in vrows:
        for nm in re.findall(r"`([A-Za-z0-9_]+)`", item):
            if nm.upper() in byname:
                valmap.setdefault(nm.upper(), []).append("[%s] %s" % (sev, rule))
    for d in (routing, universe, valmap):
        for k in d:
            d[k] = list(dict.fromkeys(d[k]))
    return routing, universe, valmap, {"rules": len(skips), "matched": matched,
                                       "unmatched": sorted(set(unmatched)),
                                       "universe_vars": len(universe),
                                       "val_rows": len(vrows), "val_vars": len(valmap)}


CONVENTIONS = [
    "Values in the datasets are RAW stored codes; this codebook is the labeling truth.",
    "Missing-value conventions (Shared Codebook v0.8 / tracker #743):",
    "  categorical items: 8 / 98 = Don't know · 9 / 99 = Refused (per value set)",
    "  amount/continuous items: -98 = Don't know · -99 = Refused (negative sentinels)",
    "  harmonized Stata output maps these to .c (DK), .b (Refused), .a (skip-logic NOTAPPL)",
    "questionnaire_number is a 12-digit string key (RR-PP-MMM-FF-CCC); breakout exports",
    "  may drop a leading region zero — restore by zero-filling to 12.",
    "Skip logic: 'Universe / applicability' = who is asked (derived from the Skip-Logic",
    "  spec); outbound routing shown on the variable that triggers it. Executable truth",
    "  stays in the instrument logic (.apc).",
    "Validation & editing: per-variable rules in 'Validation rules' + the Validations",
    "  sheet (HARD = block save, SOFT = warn-and-confirm, GATE = display-only) — maps to",
    "  DDI's data-editing/cleanOps element at dissemination.",
    "DDI mapping: Variable→var/@name · Label→var/labl · Question→var/qstn ·",
    "  Universe→var/universe · Values→var/catgry · Special codes→missing ·",
    "  packaged as DDI 2.5 at dissemination.",
]
HEADERS = ["Variable", "Label", "Question (EN, literal)", "Universe / applicability",
           "Skip / routing (outbound)", "Record / Roster", "Type",
           "Values & categories", "Special codes (DK/Refused)",
           "Validation rules (CAPI-enforced)"]
WIDTHS = [24, 34, 46, 36, 36, 20, 11, 40, 24, 42]


def cover_sheet(ws, inst, ver, stamp, n_vars, src_lines):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    rows = [
        ("UHC Survey Year 2 — Official Codebook", 16, True, DARKGREEN),
        (NAMES[inst], 14, True, GREEN),
        ("", None, False, None),
        ("Version: %s" % ver, 11, True, None),
        ("Generated: %s UTC — regenerated from the build artifacts every release; "
         "never hand-edited." % stamp, 10, False, None),
        ("Sources: " + " · ".join(src_lines), 10, False, None),
        ("", None, False, None),
        ("Variables documented: %d" % n_vars, 11, True, None),
        ("", None, False, None),
        ("Conventions", 12, True, GREEN),
    ] + [(c, 10, False, None) for c in CONVENTIONS]
    for i, (text, size, bold, color) in enumerate(rows, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(size=size or 10, bold=bold, color=color or "1C2B25")
        c.alignment = Alignment(wrap_text=True, vertical="top")


def row_cells(it, qsf, universe, routing, valmap):
    u = it["uname"]
    return [it["name"], it["label"], qsf.get(u, ""),
            "\n".join(universe.get(u, [])), "\n".join(routing.get(u, [])),
            it["record"], ("%s %s" % (it["type"], it["width"])).strip(),
            "\n".join(it["values"]), "\n".join(it["specials"]),
            "\n".join(valmap.get(u, []))]


def variables_sheet(ws, rows, qsf, universe=None, routing=None, valmap=None):
    universe, routing, valmap = universe or {}, routing or {}, valmap or {}
    ws.freeze_panes = "A2"
    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=GREEN)
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    r = 2
    for it in rows:
        vals = row_cells(it, qsf, universe, routing, valmap)
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(size=9, name="Consolas" if col == 1 else "Calibri")
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(col != 1))
        if r % 2 == 0:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=INKLIGHT)
        r += 1


def validations_sheet(ws, vrows):
    for col, (h, w) in enumerate((("Item(s)", 44), ("Rule", 90), ("Severity", 20)), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=GREEN)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    for r, (item, rule, sev) in enumerate(vrows, start=2):
        for col, v in enumerate((item.replace("`", ""), rule, sev), start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(size=9, name="Consolas" if col == 1 else "Calibri")
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(col != 1))


def build_instrument_xlsx(inst, out_dir, versions, stamp, cspro):
    folder, base = BASE[inst]
    rows = parse_dcf(os.path.join(cspro, folder, base + ".dcf"))
    qsf = load_qsf(os.path.join(cspro, folder, base + ".ent.qsf"))
    logic = os.path.join(cspro, folder, "%s-Skip-Logic-and-Validations.md" % folder)
    skips, vrows = parse_logic_doc(logic) if os.path.exists(logic) else ([], [])
    routing, universe, valmap, cov = logic_maps(skips, vrows, rows)
    ver = versions.get(folder, {})
    vlabel = "v%s (%s)" % (ver.get("version", "?"), ver.get("date", "?"))
    wb = Workbook()
    cover_sheet(wb.active, inst, vlabel, stamp, len(rows),
                ["%s.dcf" % base, "%s.ent.qsf (EN)" % base,
                 "%s-Skip-Logic-and-Validations.md" % folder, "versions.json"])
    wb.active.title = "Cover"
    variables_sheet(wb.create_sheet("Variables"), rows, qsf, universe, routing, valmap)
    if vrows:
        validations_sheet(wb.create_sheet("Validations"), vrows)
    fn = "%s-Codebook-v%s.xlsx" % (folder, ver.get("version", "x"))
    wb.save(os.path.join(out_dir, fn))
    return fn, rows, qsf, vlabel, universe, routing, valmap, vrows, cov


def build_f2_xlsx(out_dir, stamp, f2labels):
    rows = []
    for k, v in f2labels.items():
        rows.append({"name": k, "label": v.get("label", ""), "record": "f2_responses",
                     "type": v.get("type", "text"), "width": "", "occ": 1,
                     "values": ["stored as English display string(s); multi-select "
                                "answers joined with '; '"] if v.get("type") not in
                               ("number",) else [],
                     "specials": [], "uname": k})
    wb = Workbook()
    cover_sheet(wb.active, "f2", "PWA (spec-versioned in-app)", stamp, len(rows),
                ["f2-item-labels.json (from the PWA generated items)",
                 "skip logic enforced in-app (shouldShow) — see F2-Spec.md"])
    wb.active.title = "Cover"
    variables_sheet(wb.create_sheet("Variables"), rows, {})
    fn = "F2-Codebook-%s.xlsx" % stamp[:10]
    wb.save(os.path.join(out_dir, fn))
    return fn, rows


CHROME = [r"C:\Program Files\Google\Chrome\Application\chrome.exe",
          r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
          "google-chrome", "chromium"]


def to_pdf(html_path):
    """Print a codebook HTML to PDF with headless Chrome/Edge (no external CSS,
    so the render is deterministic). Returns the pdf path, or None if no browser."""
    import subprocess
    exe = next((p for p in CHROME if os.path.exists(p) or "/" not in p and
                not p.endswith(".exe")), None)
    exe = next((p for p in CHROME if os.path.exists(p)), exe)
    if not exe:
        return None
    pdf = html_path[:-5] + ".pdf"
    r = subprocess.run([exe, "--headless=new", "--disable-gpu", "--no-sandbox",
                        "--no-pdf-header-footer", "--print-to-pdf-no-header",
                        "--print-to-pdf=" + pdf, "file:///" + html_path.replace("\\", "/")],
                       capture_output=True, text=True, timeout=600)
    return pdf if os.path.exists(pdf) else None


HTML_TOP = """<!doctype html><html lang="en"><head><meta charset="utf-8">
<title>UHC Survey Year 2 — Official Codebook</title><style>
body{font:13px/1.45 Calibri,system-ui,sans-serif;color:#1c2b25;margin:28px}
h1{color:#004d2c;font-size:22px;margin:0}
h2{color:#006b3f;font-size:17px;border-bottom:2px solid #006b3f;padding-bottom:4px;margin:34px 0 10px;page-break-before:always}
h2:first-of-type{page-break-before:auto}
.meta{color:#5b6b63;font-size:12px;margin:6px 0 18px}
.conv{background:#f4f7f5;border:1px solid #dfe7e2;border-radius:8px;padding:10px 14px;font-size:12px;white-space:pre-line}
table{border-collapse:collapse;width:100%;font-size:11px;margin-top:8px}
th{background:#006b3f;color:#fff;text-align:left;padding:5px 7px}
td{border:1px solid #b7c4bd;padding:4px 7px;vertical-align:top}
tr:nth-child(even) td{background:#eef3f0}
td.v{font-family:Consolas,monospace;white-space:nowrap}
td.c{white-space:pre-line}
@media print{h2{page-break-before:always}}
</style></head><body>
"""


def html_section(inst, vlabel, rows, qsf, universe=None, routing=None, valmap=None,
                 vrows=None):
    universe, routing, valmap = universe or {}, routing or {}, valmap or {}
    out = ["<h2>%s — %s</h2>" % (html.escape(NAMES[inst]), html.escape(vlabel))]
    out.append("<table><tr>%s</tr>" % "".join("<th>%s</th>" % h for h in HEADERS))
    for it in rows:
        cells = row_cells(it, qsf, universe, routing, valmap)
        out.append("<tr>" + "".join(
            '<td class="%s">%s</td>' % ("v" if i == 0 else "c", html.escape(str(c)))
            for i, c in enumerate(cells)) + "</tr>")
    out.append("</table>")
    if vrows:
        out.append("<h3>Validations (CAPI-enforced; HARD = block, SOFT = warn, "
                   "GATE = display-only)</h3>")
        out.append("<table><tr><th>Item(s)</th><th>Rule</th><th>Severity</th></tr>")
        for item, rule, sev in vrows:
            out.append("<tr><td class=\"v\">%s</td><td class=\"c\">%s</td><td>%s</td></tr>"
                       % (html.escape(item.replace("`", "")), html.escape(rule),
                          html.escape(sev)))
        out.append("</table>")
    return "\n".join(out)


def main():
    ap = argparse.ArgumentParser(description="Generate the official UHC-Y2 codebooks.")
    ap.add_argument("--out-dir", default=os.path.join(HERE, "codebook-out"))
    ap.add_argument("--cspro-dir", default=CSPRO,
                    help="deliverables/CSPro to document — point at the MAIN checkout "
                         "(the deploy truth); a worktree copy may lag the deployed "
                         "versions (default: %(default)s)")
    ap.add_argument("--no-pdf", action="store_true",
                    help="skip the headless-browser PDF pass (xlsx + html only)")
    a = ap.parse_args()
    os.makedirs(a.out_dir, exist_ok=True)
    versions = json.load(open(os.path.join(a.cspro_dir, "versions.json"), encoding="utf-8"))
    # f2-item-labels.json ships with the CSWeb generators; until Carl merges the
    # f2-productivity-panel branch it exists only in that worktree — search both.
    import glob
    repo = os.path.abspath(os.path.join(DELIV, ".."))
    candidates = [os.path.join(DELIV, "CSWeb", "f2-item-labels.json")] + sorted(glob.glob(
        os.path.join(repo, ".claude", "worktrees", "*", "deliverables", "CSWeb",
                     "f2-item-labels.json")))
    f2path = next((p for p in candidates if os.path.exists(p)), None)
    if not f2path:
        sys.exit("f2-item-labels.json not found (looked in %d places) — regenerate it "
                 "from the PWA (see csweb-spss-gen.py footer note)." % len(candidates))
    f2labels = json.load(open(f2path, encoding="utf-8"))
    stamp = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M")

    sections, made, covs = [], [], {}
    manifest = {"generated": stamp + " UTC", "instruments": {}}

    def page(title_html, body_html):
        return (HTML_TOP + "<h1>UHC Survey Year 2 — Official Codebook</h1>\n"
                + '<div class="meta">%s</div>\n' % title_html
                + '<div class="conv">%s</div>\n' % html.escape("\n".join(CONVENTIONS))
                + body_html + "</body></html>\n")

    meta_line = ("Generated %s UTC from the build artifacts (dictionary + question text + "
                 "skip-logic/validation spec) — regenerate, never hand-edit. Variable "
                 "documentation follows the DDI-Codebook 2.5 element set used by the PSA "
                 "Data Archive (PSADA/NADA); presentation follows the DHS recode-manual "
                 "table style." % stamp)

    for inst in ("f1", "f3", "f4"):
        (fn, rows, qsf, vlabel, universe, routing, valmap,
         vrows, cov) = build_instrument_xlsx(inst, a.out_dir, versions, stamp,
                                             a.cspro_dir)
        made.append((fn, len(rows)))
        covs[inst] = cov
        sec = html_section(inst, vlabel, rows, qsf, universe, routing, valmap, vrows)
        sections.append(sec)
        folder = BASE[inst][0]
        hp = os.path.join(a.out_dir, "%s-Codebook-v%s.html"
                          % (folder, versions.get(folder, {}).get("version", "x")))
        with open(hp, "w", encoding="utf-8") as f:
            f.write(page(meta_line, sec))
        pdf = to_pdf(hp) if not a.no_pdf else None
        manifest["instruments"][inst] = {
            "name": NAMES[inst], "version": versions.get(folder, {}).get("version", ""),
            "date": versions.get(folder, {}).get("date", ""), "variables": len(rows),
            "xlsx": fn, "pdf": os.path.basename(pdf) if pdf else "",
            "skip_rules": cov["rules"], "validations": cov["val_rows"]}
    fn2, rows2 = build_f2_xlsx(a.out_dir, stamp, f2labels)
    made.append((fn2, len(rows2)))
    sec2 = html_section("f2", "PWA", rows2, {})
    sections.append(sec2)
    hp2 = os.path.join(a.out_dir, "F2-Codebook-%s.html" % stamp[:10])
    with open(hp2, "w", encoding="utf-8") as f:
        f.write(page(meta_line, sec2))
    pdf2 = to_pdf(hp2) if not a.no_pdf else None
    manifest["instruments"]["f2"] = {
        "name": NAMES["f2"], "version": "PWA", "date": stamp[:10],
        "variables": len(rows2), "xlsx": fn2,
        "pdf": os.path.basename(pdf2) if pdf2 else "", "skip_rules": 0, "validations": 0}

    hname = "UHC-Y2-Codebook-%s.html" % stamp[:10]
    hpath = os.path.join(a.out_dir, hname)
    with open(hpath, "w", encoding="utf-8") as f:
        f.write(page(meta_line, "\n".join(sections)))
    combined_pdf = to_pdf(hpath) if not a.no_pdf else None
    manifest["combined"] = {"html": hname,
                            "pdf": os.path.basename(combined_pdf) if combined_pdf else "",
                            "zip": "uhc-year2-codebooks.zip"}
    # one zip carrying every format, for the data room's single-click download
    import zipfile
    zpath = os.path.join(a.out_dir, "uhc-year2-codebooks.zip")
    with zipfile.ZipFile(zpath, "w", zipfile.ZIP_DEFLATED) as z:
        for fn in sorted(os.listdir(a.out_dir)):
            if fn.endswith((".xlsx", ".pdf")) or fn == hname:
                z.write(os.path.join(a.out_dir, fn), fn)
    with open(os.path.join(a.out_dir, "codebook-manifest.json"), "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=1)

    print("codebooks -> %s" % a.out_dir)
    for fn, n in made:
        print("  %-28s %4d variables" % (fn, n))
    print("  %s (all instruments, printable)" % hname)
    print("logic coverage (skip/validation join):")
    for inst, c in covs.items():
        print("  %-4s %d skip rules (%d source-matched), universe on %d vars, "
              "%d validation rows (%d vars annotated)"
              % (inst, c["rules"], c["matched"], c["universe_vars"],
                 c["val_rows"], c["val_vars"]))
        if c["unmatched"]:
            print("       unmatched skip sources: %s" % ", ".join(c["unmatched"]))


if __name__ == "__main__":
    main()
